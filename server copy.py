from fastapi import FastAPI, APIRouter, HTTPException, status, Depends, UploadFile, File, Response
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
from fastapi import APIRouter, HTTPException, Depends, Body
import socket
import os
import logging
from pathlib import Path
from pydantic import BaseModel, Field, ConfigDict
from typing import List, Optional
from enum import Enum
import uuid
from datetime import datetime, timezone, timedelta
import bcrypt
import jwt
import barcode
from barcode.writer import ImageWriter
from io import BytesIO
import base64
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

app = FastAPI()
api_router = APIRouter(prefix="/api")
security = HTTPBearer()

JWT_SECRET = os.environ.get('JWT_SECRET', 'your-secret-key-change-in-production')

# ─────────────────────────────────────────────────────────────
#  ROLES
#  admin   → full access (all CRUD, users, reports, delete)
#  billing → POS sales + view own orders only
#  manager → inventory, stock, locations, racks, transfers, daybook, reports
#  staff   → read-only on products, inventory, daybook, locations
# ─────────────────────────────────────────────────────────────

VALID_ROLES = ["admin", "billing", "manager", "staff"]

def require_roles(user, allowed: list, detail: str = "Access denied"):
    """Raise 403 if user.role is not in allowed list."""
    if user.role not in allowed:
        raise HTTPException(status_code=403, detail=detail)


class UserRole(str, Enum):
    ADMIN   = "admin"
    BILLING = "billing"
    MANAGER = "manager"
    STAFF   = "staff"


# ─── Models ───────────────────────────────────────────────────

class DayBookEntry(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    type: str
    amount: float
    party_name: str
    party_type: str
    assigned_to_user_id: Optional[str] = None
    assigned_to_name: Optional[str] = None
    notes: Optional[str] = ""
    date: str
    status: str = "pending"
    created_by: str
    created_by_name: str
    settled_at: Optional[str] = None
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))


class DayBookEntryCreate(BaseModel):
    type: str
    amount: float
    party_name: str
    party_type: str
    assigned_to_user_id: Optional[str] = None
    notes: Optional[str] = ""
    date: str


class User(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    email: str
    name: str
    role: str = "staff"
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))


class UserCreate(BaseModel):
    email: str
    password: str
    name: str
    role: Optional[str] = "staff"


class UserLogin(BaseModel):
    email: str
    password: str


class Product(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    name: str
    sku: str
    barcode: str
    category: str
    purchase_price: float
    selling_price: float
    stock_quantity: int
    supplier: Optional[str] = ""
    image_url: Optional[str] = ""
    low_stock_threshold: int = 10
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))


class ProductCreate(BaseModel):
    name: str
    sku: str
    category: str
    purchase_price: float
    selling_price: float
    stock_quantity: int
    supplier: Optional[str] = ""
    image_url: Optional[str] = ""
    low_stock_threshold: Optional[int] = 10


class ProductUpdate(BaseModel):
    name: Optional[str] = None
    sku: Optional[str] = None
    category: Optional[str] = None
    purchase_price: Optional[float] = None
    selling_price: Optional[float] = None
    stock_quantity: Optional[int] = None
    supplier: Optional[str] = None
    image_url: Optional[str] = None
    low_stock_threshold: Optional[int] = None


class OrderItem(BaseModel):
    product_id: str
    product_name: str
    quantity: int
    price: float
    total: float
    # GST fields (optional — POS sends these, old orders won't have them)
    hsn: Optional[str] = ""
    unit: Optional[str] = "Nos"
    discount: Optional[float] = 0
    gst_rate: Optional[float] = 18
    cgst: Optional[float] = 0
    sgst: Optional[float] = 0
    igst: Optional[float] = 0
    taxable: Optional[float] = 0


class Order(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    invoice_number: str
    items: List[OrderItem]
    subtotal: float
    tax: float
    discount: float
    total: float
    # Extended GST fields
    cgst: Optional[float] = 0
    sgst: Optional[float] = 0
    igst: Optional[float] = 0
    total_tax: Optional[float] = 0
    round_off: Optional[float] = 0
    customer_name: Optional[str] = ""
    salesman_id: Optional[str] = ""
    narration: Optional[str] = ""
    payment_mode: Optional[str] = "Cash"
    gst_type: Optional[str] = "CGST+SGST"
    invoice_date: Optional[str] = ""
    created_by: str
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))


class OrderCreate(BaseModel):
    items: List[OrderItem]
    subtotal: float
    tax: float
    discount: float
    total: float
    cgst: Optional[float] = 0
    sgst: Optional[float] = 0
    igst: Optional[float] = 0
    total_tax: Optional[float] = 0
    round_off: Optional[float] = 0
    customer_name: Optional[str] = ""
    salesman_id: Optional[str] = ""
    narration: Optional[str] = ""
    payment_mode: Optional[str] = "Cash"
    gst_type: Optional[str] = "CGST+SGST"
    invoice_date: Optional[str] = ""


class Location(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    name: str
    type: str
    description: Optional[str] = ""
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))


class LocationCreate(BaseModel):
    name: str
    type: str
    description: Optional[str] = ""


class Rack(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    code: str
    name: str
    location_id: str
    location_name: str
    description: Optional[str] = ""
    max_capacity: Optional[int] = None
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))


class RackCreate(BaseModel):
    code: str
    name: str
    location_id: str
    description: Optional[str] = ""
    max_capacity: Optional[int] = None


class RackUpdate(BaseModel):
    code: Optional[str] = None
    name: Optional[str] = None
    location_id: Optional[str] = None
    description: Optional[str] = None
    max_capacity: Optional[int] = None


class RackAssignment(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    product_id: str
    product_name: str
    rack_id: str
    rack_code: str
    location_id: str
    location_name: str
    quantity: int
    assigned_by: str
    assigned_by_name: str
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    updated_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))


class RackAssignmentCreate(BaseModel):
    product_id: str
    rack_id: str
    quantity: int


class RackAssignmentUpdate(BaseModel):
    quantity: int


class StockTransfer(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    product_id: str
    product_name: str
    from_rack_id: Optional[str] = None
    from_rack_code: Optional[str] = None
    to_rack_id: Optional[str] = None
    to_rack_code: Optional[str] = None
    quantity: int
    transfer_type: str
    notes: Optional[str] = ""
    transferred_by: str
    transferred_by_name: str
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))


class StockTransferCreate(BaseModel):
    product_id: str
    from_rack_id: Optional[str] = None
    to_rack_id: Optional[str] = None
    quantity: int
    transfer_type: str = 'manual'
    notes: Optional[str] = ""


class ProductLocationThreshold(BaseModel):
    product_id: str
    mall_threshold: int = 5
    warehouse_threshold: int = 20


# ─── Auth Helpers ─────────────────────────────────────────────

def hash_password(password: str) -> str:
    return bcrypt.hashpw(password.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')


def verify_password(password: str, hashed: str) -> bool:
    return bcrypt.checkpw(password.encode('utf-8'), hashed.encode('utf-8'))


def create_token(user_id: str, email: str, role: str) -> str:
    payload = {
        'user_id': user_id,
        'email': email,
        'role': role,
        'exp': datetime.now(timezone.utc) + timedelta(days=7)
    }
    return jwt.encode(payload, JWT_SECRET, algorithm='HS256')


async def get_current_user(credentials: HTTPAuthorizationCredentials = Depends(security)):
    try:
        token = credentials.credentials
        payload = jwt.decode(token, JWT_SECRET, algorithms=['HS256'])
        user = await db.users.find_one({'id': payload['user_id']}, {'_id': 0})
        if not user:
            raise HTTPException(status_code=401, detail="User not found")
        return User(**user)
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Token expired")
    except Exception:
        raise HTTPException(status_code=401, detail="Invalid token")


def generate_barcode_image(code: str) -> str:
    EAN = barcode.get_barcode_class('code128')
    ean = EAN(code, writer=ImageWriter())
    buffer = BytesIO()
    ean.write(buffer)
    buffer.seek(0)
    return base64.b64encode(buffer.getvalue()).decode('utf-8')


@api_router.post("/print-thermal")
async def print_thermal(
    data: dict = Body(...),
    current_user: User = Depends(get_current_user)
):
    """
    Send TSPL commands to a TSC TE144 (or any TSC/Zebra) thermal label
    printer over TCP/IP (raw port 9100).
 
    Required .env keys:
        THERMAL_PRINTER_IP   = 192.168.x.x
        THERMAL_PRINTER_PORT = 9100          (default)
 
    ENCODING NOTE:
        TSC TE144 firmware uses Latin-1 (cp1252).
        We encode with  errors="replace"  so the ₹ rupee sign (U+20B9),
        which has no cp1252 equivalent, becomes "?" without crashing.
        The frontend already substitutes "Rs." for ₹ in TSPL TEXT commands,
        so this is only a safety net.
 
    TIMEOUT:
        5 s connect + 10 s send — adequate for a local LAN printer.
        Increase THERMAL_PRINTER_TIMEOUT in .env for slower networks.
    """
    PRINTER_IP   = os.environ.get("THERMAL_PRINTER_IP", "").strip()
    PRINTER_PORT = int(os.environ.get("THERMAL_PRINTER_PORT", "9100"))
    TIMEOUT      = float(os.environ.get("THERMAL_PRINTER_TIMEOUT", "10"))
 
    if not PRINTER_IP:
        raise HTTPException(
            status_code=503,
            detail=(
                "Thermal printer not configured. "
                "Add THERMAL_PRINTER_IP to your .env file "
                "(e.g.  THERMAL_PRINTER_IP=192.168.1.100)."
            ),
        )
 
    commands: str = data.get("commands", "").strip()
    if not commands:
        raise HTTPException(status_code=400, detail="No TSPL commands provided.")
 
    # ── Encode for TSC firmware ───────────────────────────────
    #  latin-1 / cp1252  — NOT utf-8
    #  errors="replace" turns any unmappable char (e.g. ₹) into "?"
    try:
        payload: bytes = commands.encode("latin-1", errors="replace")
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"TSPL encoding error: {exc}")
 
    # ── TCP send ──────────────────────────────────────────────
    sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
    sock.settimeout(TIMEOUT)
    try:
        sock.connect((PRINTER_IP, PRINTER_PORT))
        # sendall() retries until all bytes are written (important for large jobs)
        sock.sendall(payload)
        return {
            "status": "ok",
            "printer": f"{PRINTER_IP}:{PRINTER_PORT}",
            "bytes_sent": len(payload),
        }
    except socket.timeout:
        raise HTTPException(
            status_code=504,
            detail=f"Printer at {PRINTER_IP}:{PRINTER_PORT} timed out after {TIMEOUT}s. "
                   "Check the printer is on and connected to the same network.",
        )
    except ConnectionRefusedError:
        raise HTTPException(
            status_code=503,
            detail=f"Connection refused by printer at {PRINTER_IP}:{PRINTER_PORT}. "
                   "Ensure port 9100 is open on the printer (Interface → TCP/IP settings).",
        )
    except OSError as exc:
        raise HTTPException(
            status_code=503,
            detail=f"Network error reaching printer at {PRINTER_IP}: {exc}",
        )
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"Unexpected print error: {exc}")
    finally:
        try:
            sock.close()
        except Exception:
            pass

# ═════════════════════════════════════════════════════════════
#  AUTH ENDPOINTS
# ═════════════════════════════════════════════════════════════

@api_router.post("/auth/register", response_model=User)
async def register(
    user_data: UserCreate,
    current_user: Optional[User] = None   # optional — first admin is created freely
):
    existing = await db.users.find_one({'email': user_data.email}, {'_id': 0})
    if existing:
        raise HTTPException(status_code=400, detail="Email already registered")

    # Validate role
    role = user_data.role if user_data.role in VALID_ROLES else "staff"

    user = User(email=user_data.email, name=user_data.name, role=role)
    doc = user.model_dump()
    doc['password'] = hash_password(user_data.password)
    doc['created_at'] = doc['created_at'].isoformat()
    await db.users.insert_one(doc)
    return user


@api_router.post("/auth/login")
async def login(credentials: UserLogin):
    user_doc = await db.users.find_one({'email': credentials.email}, {'_id': 0})
    if not user_doc or not verify_password(credentials.password, user_doc['password']):
        raise HTTPException(status_code=401, detail="Invalid credentials")

    token = create_token(user_doc['id'], user_doc['email'], user_doc['role'])
    return {
        'token': token,
        'user': {
            'id':    user_doc['id'],
            'email': user_doc['email'],
            'name':  user_doc['name'],
            'role':  user_doc['role'],
        }
    }


@api_router.get("/auth/me", response_model=User)
async def get_me(current_user: User = Depends(get_current_user)):
    return current_user


# ═════════════════════════════════════════════════════════════
#  PRODUCTS
#  GET    → all roles
#  POST   → admin, manager
#  PUT    → admin, manager
#  DELETE → admin only
# ═════════════════════════════════════════════════════════════

@api_router.get("/products", response_model=List[Product])
async def get_products(current_user: User = Depends(get_current_user)):
    # All roles can view products
    products = await db.products.find({}, {'_id': 0}).to_list(1000)
    for p in products:
        if isinstance(p.get('created_at'), str):
            p['created_at'] = datetime.fromisoformat(p['created_at'])
    return products


@api_router.post("/products", response_model=Product)
async def create_product(
    product_data: ProductCreate,
    current_user: User = Depends(get_current_user)
):
    require_roles(current_user, ['admin', 'manager'],
                  "Only admins and managers can create products")

    # Use SKU as the barcode so printed labels and scanner always match
    barcode_num = product_data.sku.strip() if product_data.sku.strip() else str(uuid.uuid4().int)[:12]
    product = Product(**product_data.model_dump(), barcode=barcode_num)
    doc = product.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    await db.products.insert_one(doc)
    return product


@api_router.get("/products/{product_id}", response_model=Product)
async def get_product(product_id: str, current_user: User = Depends(get_current_user)):
    product = await db.products.find_one({'id': product_id}, {'_id': 0})
    if not product:
        raise HTTPException(status_code=404, detail="Product not found")
    if isinstance(product.get('created_at'), str):
        product['created_at'] = datetime.fromisoformat(product['created_at'])
    return Product(**product)


@api_router.get("/products/barcode/{barcode_num}")
async def get_product_by_barcode(barcode_num: str, current_user: User = Depends(get_current_user)):
    # Search by barcode field first, then fall back to SKU field
    # This allows scanning both auto-generated barcodes AND manually entered SKU codes (e.g. from Vyapar)
    product = await db.products.find_one(
        {'$or': [{'barcode': barcode_num}, {'sku': barcode_num}]},
        {'_id': 0}
    )
    if not product:
        raise HTTPException(status_code=404, detail="Product not found")
    if isinstance(product.get('created_at'), str):
        product['created_at'] = datetime.fromisoformat(product['created_at'])
    return product


@api_router.put("/products/{product_id}", response_model=Product)
async def update_product(
    product_id: str,
    update_data: ProductUpdate,
    current_user: User = Depends(get_current_user)
):
    require_roles(current_user, ['admin', 'manager'],
                  "Only admins and managers can update products")

    existing = await db.products.find_one({'id': product_id}, {'_id': 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Product not found")

    update_dict = {k: v for k, v in update_data.model_dump().items() if v is not None}
    if update_dict:
        await db.products.update_one({'id': product_id}, {'$set': update_dict})

    updated = await db.products.find_one({'id': product_id}, {'_id': 0})
    if isinstance(updated.get('created_at'), str):
        updated['created_at'] = datetime.fromisoformat(updated['created_at'])
    return Product(**updated)


@api_router.delete("/products/{product_id}")
async def delete_product(product_id: str, current_user: User = Depends(get_current_user)):
    require_roles(current_user, ['admin'], "Only admins can delete products")

    result = await db.products.delete_one({'id': product_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Product not found")
    return {'message': 'Product deleted successfully'}


@api_router.get("/products/{product_id}/barcode-image")
async def get_barcode_image(product_id: str, current_user: User = Depends(get_current_user)):
    product = await db.products.find_one({'id': product_id}, {'_id': 0})
    if not product:
        raise HTTPException(status_code=404, detail="Product not found")
    barcode_img = generate_barcode_image(product['barcode'])
    return {'barcode_image': f'data:image/png;base64,{barcode_img}'}


@api_router.get("/barcode-preview/{code}")
async def preview_barcode_image(code: str, current_user: User = Depends(get_current_user)):
    """Generate a barcode image for ANY code — used for live preview before product is saved."""
    try:
        barcode_img = generate_barcode_image(code.strip())
        return {'barcode_image': f'data:image/png;base64,{barcode_img}'}
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Cannot generate barcode for this code: {str(e)}")


# ═════════════════════════════════════════════════════════════
#  ORDERS / POS
#  POST (checkout) → admin, billing
#  GET (list)      → admin, billing (own only), manager (view all)
# ═════════════════════════════════════════════════════════════

@api_router.post("/orders", response_model=Order)
async def create_order(
    order_data: OrderCreate,
    current_user: User = Depends(get_current_user)
):
    require_roles(current_user, ['admin', 'billing'],
                  "Only admins and billing staff can create orders")

    invoice_num = f"INV-{datetime.now().strftime('%Y%m%d')}-{str(uuid.uuid4())[:8].upper()}"

    order = Order(
        invoice_number=invoice_num,
        items=order_data.items,
        subtotal=order_data.subtotal,
        tax=order_data.tax,
        discount=order_data.discount,
        total=order_data.total,
        cgst=order_data.cgst,
        sgst=order_data.sgst,
        igst=order_data.igst,
        total_tax=order_data.total_tax,
        round_off=order_data.round_off,
        customer_name=order_data.customer_name,
        salesman_id=order_data.salesman_id,
        narration=order_data.narration,
        payment_mode=order_data.payment_mode,
        gst_type=order_data.gst_type,
        invoice_date=order_data.invoice_date,
        created_by=current_user.id
    )

    # Deduct stock from mall rack assignments
    mall_locations = await db.locations.find({'type': 'mall'}, {'_id': 0}).to_list(10)
    mall_location_ids = [loc['id'] for loc in mall_locations]

    for item in order_data.items:
        product = await db.products.find_one({'id': item.product_id}, {'_id': 0})
        if product:
            new_stock = product['stock_quantity'] - item.quantity
            await db.products.update_one(
                {'id': item.product_id},
                {'$set': {'stock_quantity': max(0, new_stock)}}
            )

            remaining_qty = item.quantity
            mall_assignments = await db.rack_assignments.find({
                'product_id': item.product_id,
                'location_id': {'$in': mall_location_ids}
            }, {'_id': 0}).sort('quantity', -1).to_list(100)

            for assignment in mall_assignments:
                if remaining_qty <= 0:
                    break
                deduct_qty = min(remaining_qty, assignment['quantity'])
                new_rack_qty = assignment['quantity'] - deduct_qty
                if new_rack_qty == 0:
                    await db.rack_assignments.delete_one({'id': assignment['id']})
                else:
                    await db.rack_assignments.update_one(
                        {'id': assignment['id']},
                        {'$set': {
                            'quantity': new_rack_qty,
                            'updated_at': datetime.now(timezone.utc).isoformat()
                        }}
                    )

                transfer = StockTransfer(
                    product_id=item.product_id,
                    product_name=product['name'],
                    from_rack_id=assignment['rack_id'],
                    from_rack_code=assignment['rack_code'],
                    quantity=deduct_qty,
                    transfer_type='sale',
                    notes=f"Sale deduction from invoice {invoice_num}",
                    transferred_by=current_user.id,
                    transferred_by_name=current_user.name
                )
                transfer_doc = transfer.model_dump()
                transfer_doc['created_at'] = transfer_doc['created_at'].isoformat()
                await db.stock_transfers.insert_one(transfer_doc)
                remaining_qty -= deduct_qty

    doc = order.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    await db.orders.insert_one(doc)
    return order


@api_router.get("/orders", response_model=List[Order])
async def get_orders(current_user: User = Depends(get_current_user)):
    require_roles(current_user, ['admin', 'billing', 'manager'],
                  "Staff cannot view orders")

    # billing staff see only their own orders
    query = {} if current_user.role in ['admin', 'manager'] else {'created_by': current_user.id}
    orders = await db.orders.find(query, {'_id': 0}).sort('created_at', -1).to_list(1000)
    for order in orders:
        if isinstance(order.get('created_at'), str):
            order['created_at'] = datetime.fromisoformat(order['created_at'])
    return orders


@api_router.get("/orders/{order_id}", response_model=Order)
async def get_order(order_id: str, current_user: User = Depends(get_current_user)):
    require_roles(current_user, ['admin', 'billing', 'manager'],
                  "Staff cannot view orders")

    order = await db.orders.find_one({'id': order_id}, {'_id': 0})
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")
    # billing can only see their own orders
    if current_user.role == 'billing' and order.get('created_by') != current_user.id:
        raise HTTPException(status_code=403, detail="Access denied")
    if isinstance(order.get('created_at'), str):
        order['created_at'] = datetime.fromisoformat(order['created_at'])
    return Order(**order)


# ═════════════════════════════════════════════════════════════
#  REPORTS
#  admin, manager only
# ═════════════════════════════════════════════════════════════

@api_router.get("/reports/sales")
async def get_sales_report(
    period: str = "daily",
    current_user: User = Depends(get_current_user)
):
    require_roles(current_user, ['admin', 'manager'],
                  "Only admins and managers can view reports")

    now = datetime.now(timezone.utc)
    if period == "daily":
        start_date = now.replace(hour=0, minute=0, second=0, microsecond=0)
    elif period == "weekly":
        start_date = now - timedelta(days=7)
    elif period == "monthly":
        start_date = now - timedelta(days=30)
    else:
        start_date = now - timedelta(days=30)

    orders = await db.orders.find({}, {'_id': 0}).to_list(10000)
    filtered_orders = []
    for order in orders:
        created_at = order.get('created_at')
        if isinstance(created_at, str):
            created_at = datetime.fromisoformat(created_at)
        if created_at >= start_date:
            filtered_orders.append(order)

    total_sales = sum(order['total'] for order in filtered_orders)
    total_orders = len(filtered_orders)

    products = await db.products.find({}, {'_id': 0}).to_list(1000)
    total_profit = 0
    for order in filtered_orders:
        for item in order['items']:
            product = next((p for p in products if p['id'] == item['product_id']), None)
            if product:
                profit = (item['price'] - product['purchase_price']) * item['quantity']
                total_profit += profit

    return {
        'period': period,
        'total_sales': round(total_sales, 2),
        'total_orders': total_orders,
        'total_profit': round(total_profit, 2),
        'orders': filtered_orders
    }


@api_router.get("/reports/export/excel")
async def export_excel_report(
    period: str = "daily",
    current_user: User = Depends(get_current_user)
):
    require_roles(current_user, ['admin', 'manager'],
                  "Only admins and managers can export reports")

    report_data = await get_sales_report(period, current_user)

    wb = Workbook()
    ws = wb.active
    ws.title = f"{period.capitalize()} Sales Report"
    ws['A1'] = f"Sales Report - {period.capitalize()}"
    ws['A1'].font = Font(size=14, bold=True)
    ws['A3'] = 'Total Sales:'
    ws['B3'] = report_data['total_sales']
    ws['A4'] = 'Total Orders:'
    ws['B4'] = report_data['total_orders']
    ws['A5'] = 'Total Profit:'
    ws['B5'] = report_data['total_profit']

    headers = ['Invoice Number', 'Date', 'Items', 'Subtotal', 'Tax', 'Discount', 'Total']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=7, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    for row_idx, order in enumerate(report_data['orders'], start=8):
        ws.cell(row=row_idx, column=1, value=order['invoice_number'])
        created_at = order['created_at']
        if isinstance(created_at, str):
            created_at = datetime.fromisoformat(created_at)
        ws.cell(row=row_idx, column=2, value=created_at.strftime('%Y-%m-%d %H:%M'))
        ws.cell(row=row_idx, column=3, value=len(order['items']))
        ws.cell(row=row_idx, column=4, value=order['subtotal'])
        ws.cell(row=row_idx, column=5, value=order['tax'])
        ws.cell(row=row_idx, column=6, value=order['discount'])
        ws.cell(row=row_idx, column=7, value=order['total'])

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return Response(
        content=buffer.getvalue(),
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename=sales_report_{period}.xlsx'}
    )


# ═════════════════════════════════════════════════════════════
#  DASHBOARD
#  All roles (financial totals hidden for billing/staff in frontend)
# ═════════════════════════════════════════════════════════════

@api_router.get("/dashboard/stats")
async def get_dashboard_stats(current_user: User = Depends(get_current_user)):
    products = await db.products.find({}, {'_id': 0}).to_list(1000)
    orders   = await db.orders.find({}, {'_id': 0}).to_list(1000)

    now         = datetime.now(timezone.utc)
    today_start = now.replace(hour=0, minute=0, second=0, microsecond=0)

    today_orders = []
    for order in orders:
        created_at = order.get('created_at')
        if isinstance(created_at, str):
            created_at = datetime.fromisoformat(created_at)
        if created_at >= today_start:
            today_orders.append(order)

    today_sales = sum(o['total'] for o in today_orders)
    total_sales = sum(o['total'] for o in orders)

    total_profit = 0
    for order in orders:
        for item in order['items']:
            product = next((p for p in products if p['id'] == item['product_id']), None)
            if product:
                profit = (item['price'] - product['purchase_price']) * item['quantity']
                total_profit += profit

    low_stock_products = [p for p in products if p['stock_quantity'] <= p.get('low_stock_threshold', 10)]

    return {
        'today_sales':        round(today_sales, 2),
        'total_sales':        round(total_sales, 2),
        'total_profit':       round(total_profit, 2),
        'total_products':     len(products),
        'total_orders':       len(orders),
        'low_stock_count':    len(low_stock_products),
        'low_stock_products': low_stock_products[:5],
        # expose role so frontend can conditionally hide financial data
        'viewer_role':        current_user.role,
    }


# ═════════════════════════════════════════════════════════════
#  USERS
#  GET, POST, PUT role → admin only
# ═════════════════════════════════════════════════════════════

@api_router.get("/users", response_model=List[User])
async def get_users(current_user: User = Depends(get_current_user)):
    require_roles(current_user, ['admin'], "Only admins can view users")

    users = await db.users.find({}, {'_id': 0, 'password': 0}).to_list(1000)
    for u in users:
        if isinstance(u.get('created_at'), str):
            u['created_at'] = datetime.fromisoformat(u['created_at'])
    return users


@api_router.put("/users/{user_id}/role")
async def update_user_role(
    user_id: str,
    role: str,
    current_user: User = Depends(get_current_user)
):
    require_roles(current_user, ['admin'], "Only admins can update roles")

    if role not in VALID_ROLES:
        raise HTTPException(
            status_code=400,
            detail=f"Invalid role '{role}'. Must be one of: {', '.join(VALID_ROLES)}"
        )

    result = await db.users.update_one({'id': user_id}, {'$set': {'role': role}})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="User not found")
    return {'message': f'Role updated to {role}'}


# ═════════════════════════════════════════════════════════════
#  LOCATIONS
#  GET  → all roles
#  POST → admin, manager
# ═════════════════════════════════════════════════════════════

@api_router.get("/locations", response_model=List[Location])
async def get_locations(current_user: User = Depends(get_current_user)):
    locations = await db.locations.find({}, {'_id': 0}).to_list(1000)
    for loc in locations:
        if isinstance(loc.get('created_at'), str):
            loc['created_at'] = datetime.fromisoformat(loc['created_at'])
    return locations


@api_router.post("/locations", response_model=Location)
async def create_location(
    location_data: LocationCreate,
    current_user: User = Depends(get_current_user)
):
    require_roles(current_user, ['admin', 'manager'],
                  "Only admins and managers can create locations")

    location = Location(**location_data.model_dump())
    doc = location.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    await db.locations.insert_one(doc)
    return location


# ═════════════════════════════════════════════════════════════
#  RACKS
#  GET    → all roles
#  POST   → admin, manager
#  PUT    → admin, manager
#  DELETE → admin only
# ═════════════════════════════════════════════════════════════

@api_router.get("/racks", response_model=List[Rack])
async def get_racks(
    location_id: Optional[str] = None,
    current_user: User = Depends(get_current_user)
):
    query = {'location_id': location_id} if location_id else {}
    racks = await db.racks.find(query, {'_id': 0}).to_list(1000)
    for rack in racks:
        if isinstance(rack.get('created_at'), str):
            rack['created_at'] = datetime.fromisoformat(rack['created_at'])
    return racks


@api_router.post("/racks", response_model=Rack)
async def create_rack(
    rack_data: RackCreate,
    current_user: User = Depends(get_current_user)
):
    require_roles(current_user, ['admin', 'manager'],
                  "Only admins and managers can create racks")

    location = await db.locations.find_one({'id': rack_data.location_id}, {'_id': 0})
    if not location:
        raise HTTPException(status_code=404, detail="Location not found")

    existing = await db.racks.find_one({
        'code': rack_data.code, 'location_id': rack_data.location_id
    }, {'_id': 0})
    if existing:
        raise HTTPException(status_code=400, detail="Rack code already exists in this location")

    rack = Rack(**rack_data.model_dump(), location_name=location['name'])
    doc = rack.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    await db.racks.insert_one(doc)
    return rack


@api_router.get("/racks/{rack_id}", response_model=Rack)
async def get_rack(rack_id: str, current_user: User = Depends(get_current_user)):
    rack = await db.racks.find_one({'id': rack_id}, {'_id': 0})
    if not rack:
        raise HTTPException(status_code=404, detail="Rack not found")
    if isinstance(rack.get('created_at'), str):
        rack['created_at'] = datetime.fromisoformat(rack['created_at'])
    return Rack(**rack)


@api_router.put("/racks/{rack_id}", response_model=Rack)
async def update_rack(
    rack_id: str,
    update_data: RackUpdate,
    current_user: User = Depends(get_current_user)
):
    require_roles(current_user, ['admin', 'manager'],
                  "Only admins and managers can update racks")

    existing = await db.racks.find_one({'id': rack_id}, {'_id': 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Rack not found")

    update_dict = {k: v for k, v in update_data.model_dump().items() if v is not None}
    if 'location_id' in update_dict:
        location = await db.locations.find_one({'id': update_dict['location_id']}, {'_id': 0})
        if location:
            update_dict['location_name'] = location['name']
    if update_dict:
        await db.racks.update_one({'id': rack_id}, {'$set': update_dict})

    updated = await db.racks.find_one({'id': rack_id}, {'_id': 0})
    if isinstance(updated.get('created_at'), str):
        updated['created_at'] = datetime.fromisoformat(updated['created_at'])
    return Rack(**updated)


@api_router.delete("/racks/{rack_id}")
async def delete_rack(rack_id: str, current_user: User = Depends(get_current_user)):
    require_roles(current_user, ['admin'], "Only admins can delete racks")

    assignments = await db.rack_assignments.find_one({'rack_id': rack_id}, {'_id': 0})
    if assignments:
        raise HTTPException(status_code=400, detail="Cannot delete rack with product assignments")

    result = await db.racks.delete_one({'id': rack_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Rack not found")
    return {'message': 'Rack deleted successfully'}


# ═════════════════════════════════════════════════════════════
#  RACK ASSIGNMENTS
#  GET    → all roles
#  POST   → admin, manager
#  PUT    → admin, manager
#  DELETE → admin, manager
# ═════════════════════════════════════════════════════════════

@api_router.get("/products/{product_id}/rack-assignments", response_model=List[RackAssignment])
async def get_product_rack_assignments(product_id: str, current_user: User = Depends(get_current_user)):
    assignments = await db.rack_assignments.find({'product_id': product_id}, {'_id': 0}).to_list(1000)
    for a in assignments:
        if isinstance(a.get('created_at'), str):
            a['created_at'] = datetime.fromisoformat(a['created_at'])
        if isinstance(a.get('updated_at'), str):
            a['updated_at'] = datetime.fromisoformat(a['updated_at'])
    return assignments


@api_router.get("/racks/{rack_id}/products")
async def get_rack_products(rack_id: str, current_user: User = Depends(get_current_user)):
    return await db.rack_assignments.find({'rack_id': rack_id}, {'_id': 0}).to_list(1000)


@api_router.post("/rack-assignments", response_model=RackAssignment)
async def create_rack_assignment(
    assignment_data: RackAssignmentCreate,
    current_user: User = Depends(get_current_user)
):
    require_roles(current_user, ['admin', 'manager'],
                  "Only admins and managers can assign products to racks")

    product = await db.products.find_one({'id': assignment_data.product_id}, {'_id': 0})
    if not product:
        raise HTTPException(status_code=404, detail="Product not found")
    rack = await db.racks.find_one({'id': assignment_data.rack_id}, {'_id': 0})
    if not rack:
        raise HTTPException(status_code=404, detail="Rack not found")

    existing = await db.rack_assignments.find_one({
        'product_id': assignment_data.product_id,
        'rack_id':    assignment_data.rack_id
    }, {'_id': 0})
    if existing:
        raise HTTPException(status_code=400, detail="Product already assigned to this rack")

    assignments = await db.rack_assignments.find({'product_id': assignment_data.product_id}, {'_id': 0}).to_list(1000)
    total_assigned = sum(a['quantity'] for a in assignments) + assignment_data.quantity
    if total_assigned > product['stock_quantity']:
        raise HTTPException(
            status_code=400,
            detail=f"Total assigned ({total_assigned}) exceeds available stock ({product['stock_quantity']})"
        )

    assignment = RackAssignment(
        product_id=assignment_data.product_id,
        product_name=product['name'],
        rack_id=assignment_data.rack_id,
        rack_code=rack['code'],
        location_id=rack['location_id'],
        location_name=rack['location_name'],
        quantity=assignment_data.quantity,
        assigned_by=current_user.id,
        assigned_by_name=current_user.name
    )
    doc = assignment.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    doc['updated_at'] = doc['updated_at'].isoformat()
    await db.rack_assignments.insert_one(doc)
    return assignment


@api_router.put("/rack-assignments/{assignment_id}", response_model=RackAssignment)
async def update_rack_assignment(
    assignment_id: str,
    update_data: RackAssignmentUpdate,
    current_user: User = Depends(get_current_user)
):
    require_roles(current_user, ['admin', 'manager'],
                  "Only admins and managers can update rack assignments")

    existing = await db.rack_assignments.find_one({'id': assignment_id}, {'_id': 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Assignment not found")

    product = await db.products.find_one({'id': existing['product_id']}, {'_id': 0})
    other_assignments = await db.rack_assignments.find({
        'product_id': existing['product_id'], 'id': {'$ne': assignment_id}
    }, {'_id': 0}).to_list(1000)
    total_assigned = sum(a['quantity'] for a in other_assignments) + update_data.quantity
    if total_assigned > product['stock_quantity']:
        raise HTTPException(
            status_code=400,
            detail=f"Total assigned ({total_assigned}) exceeds available stock ({product['stock_quantity']})"
        )

    await db.rack_assignments.update_one(
        {'id': assignment_id},
        {'$set': {'quantity': update_data.quantity, 'updated_at': datetime.now(timezone.utc).isoformat()}}
    )
    updated = await db.rack_assignments.find_one({'id': assignment_id}, {'_id': 0})
    if isinstance(updated.get('created_at'), str):
        updated['created_at'] = datetime.fromisoformat(updated['created_at'])
    if isinstance(updated.get('updated_at'), str):
        updated['updated_at'] = datetime.fromisoformat(updated['updated_at'])
    return RackAssignment(**updated)


@api_router.delete("/rack-assignments/{assignment_id}")
async def delete_rack_assignment(assignment_id: str, current_user: User = Depends(get_current_user)):
    require_roles(current_user, ['admin', 'manager'],
                  "Only admins and managers can remove rack assignments")

    result = await db.rack_assignments.delete_one({'id': assignment_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Assignment not found")
    return {'message': 'Assignment deleted successfully'}


# ═════════════════════════════════════════════════════════════
#  STOCK TRANSFERS
#  GET  → all roles
#  POST → admin, manager
# ═════════════════════════════════════════════════════════════

@api_router.get("/stock-transfers", response_model=List[StockTransfer])
async def get_stock_transfers(
    product_id:  Optional[str] = None,
    start_date:  Optional[str] = None,
    end_date:    Optional[str] = None,
    current_user: User = Depends(get_current_user)
):
    query = {}
    if product_id:
        query['product_id'] = product_id

    transfers = await db.stock_transfers.find(query, {'_id': 0}).sort('created_at', -1).to_list(1000)

    if start_date or end_date:
        filtered = []
        for t in transfers:
            created_at = t.get('created_at')
            if isinstance(created_at, str):
                created_at = datetime.fromisoformat(created_at)
            if start_date and created_at < datetime.fromisoformat(start_date):
                continue
            if end_date and created_at > datetime.fromisoformat(end_date):
                continue
            filtered.append(t)
        transfers = filtered

    for t in transfers:
        if isinstance(t.get('created_at'), str):
            t['created_at'] = datetime.fromisoformat(t['created_at'])
    return transfers


@api_router.post("/stock-transfers", response_model=StockTransfer)
async def create_stock_transfer(
    transfer_data: StockTransferCreate,
    current_user: User = Depends(get_current_user)
):
    require_roles(current_user, ['admin', 'manager'],
                  "Only admins and managers can create stock transfers")

    product = await db.products.find_one({'id': transfer_data.product_id}, {'_id': 0})
    if not product:
        raise HTTPException(status_code=404, detail="Product not found")

    from_rack_code = None
    to_rack_code   = None

    if transfer_data.from_rack_id:
        from_assignment = await db.rack_assignments.find_one({
            'product_id': transfer_data.product_id, 'rack_id': transfer_data.from_rack_id
        }, {'_id': 0})
        if not from_assignment:
            raise HTTPException(status_code=404, detail="Product not found in source rack")
        if from_assignment['quantity'] < transfer_data.quantity:
            raise HTTPException(
                status_code=400,
                detail=f"Insufficient quantity. Available: {from_assignment['quantity']}"
            )
        new_quantity = from_assignment['quantity'] - transfer_data.quantity
        if new_quantity == 0:
            await db.rack_assignments.delete_one({'id': from_assignment['id']})
        else:
            await db.rack_assignments.update_one(
                {'id': from_assignment['id']},
                {'$set': {'quantity': new_quantity, 'updated_at': datetime.now(timezone.utc).isoformat()}}
            )
        from_rack = await db.racks.find_one({'id': transfer_data.from_rack_id}, {'_id': 0})
        from_rack_code = from_rack['code'] if from_rack else None

    if transfer_data.to_rack_id:
        to_assignment = await db.rack_assignments.find_one({
            'product_id': transfer_data.product_id, 'rack_id': transfer_data.to_rack_id
        }, {'_id': 0})
        to_rack = await db.racks.find_one({'id': transfer_data.to_rack_id}, {'_id': 0})
        to_rack_code = to_rack['code'] if to_rack else None

        if to_assignment:
            await db.rack_assignments.update_one(
                {'id': to_assignment['id']},
                {'$set': {'quantity': to_assignment['quantity'] + transfer_data.quantity,
                          'updated_at': datetime.now(timezone.utc).isoformat()}}
            )
        elif to_rack:
            assignment = RackAssignment(
                product_id=transfer_data.product_id,
                product_name=product['name'],
                rack_id=transfer_data.to_rack_id,
                rack_code=to_rack['code'],
                location_id=to_rack['location_id'],
                location_name=to_rack['location_name'],
                quantity=transfer_data.quantity,
                assigned_by=current_user.id,
                assigned_by_name=current_user.name
            )
            doc = assignment.model_dump()
            doc['created_at'] = doc['created_at'].isoformat()
            doc['updated_at'] = doc['updated_at'].isoformat()
            await db.rack_assignments.insert_one(doc)

    transfer = StockTransfer(
        product_id=transfer_data.product_id,
        product_name=product['name'],
        from_rack_id=transfer_data.from_rack_id,
        from_rack_code=from_rack_code,
        to_rack_id=transfer_data.to_rack_id,
        to_rack_code=to_rack_code,
        quantity=transfer_data.quantity,
        transfer_type=transfer_data.transfer_type,
        notes=transfer_data.notes,
        transferred_by=current_user.id,
        transferred_by_name=current_user.name
    )
    doc = transfer.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    await db.stock_transfers.insert_one(doc)
    return transfer


@api_router.get("/products/{product_id}/locations")
async def get_product_locations(product_id: str, current_user: User = Depends(get_current_user)):
    product = await db.products.find_one({'id': product_id}, {'_id': 0})
    if not product:
        raise HTTPException(status_code=404, detail="Product not found")

    assignments = await db.rack_assignments.find({'product_id': product_id}, {'_id': 0}).to_list(1000)
    mall_racks, warehouse_racks = [], []
    for a in assignments:
        location = await db.locations.find_one({'id': a['location_id']}, {'_id': 0})
        if location:
            if location['type'] == 'mall':
                mall_racks.append(a)
            else:
                warehouse_racks.append(a)

    mall_total      = sum(r['quantity'] for r in mall_racks)
    warehouse_total = sum(r['quantity'] for r in warehouse_racks)
    return {
        'product': product,
        'mall_racks': mall_racks,
        'warehouse_racks': warehouse_racks,
        'mall_total': mall_total,
        'warehouse_total': warehouse_total,
        'total_assigned': mall_total + warehouse_total
    }


@api_router.get("/dashboard/low-stock-by-location")
async def get_low_stock_by_location(current_user: User = Depends(get_current_user)):
    products  = await db.products.find({}, {'_id': 0}).to_list(1000)
    locations = await db.locations.find({}, {'_id': 0}).to_list(10)

    mall_low_stock, warehouse_low_stock = [], []
    for product in products:
        assignments = await db.rack_assignments.find({'product_id': product['id']}, {'_id': 0}).to_list(1000)
        mall_qty = warehouse_qty = 0
        for a in assignments:
            location = next((l for l in locations if l['id'] == a['location_id']), None)
            if location:
                if location['type'] == 'mall':
                    mall_qty += a['quantity']
                else:
                    warehouse_qty += a['quantity']

        if 0 < mall_qty <= 5:
            mall_low_stock.append({**product, 'current_quantity': mall_qty, 'threshold': 5})
        if 0 < warehouse_qty <= 20:
            warehouse_low_stock.append({**product, 'current_quantity': warehouse_qty, 'threshold': 20})

    return {
        'mall_low_stock':      mall_low_stock[:10],
        'warehouse_low_stock': warehouse_low_stock[:10],
        'mall_count':          len(mall_low_stock),
        'warehouse_count':     len(warehouse_low_stock)
    }


# ═════════════════════════════════════════════════════════════
#  DAY BOOK
#  GET         → admin, manager, staff (staff: own only)
#  POST        → admin, manager
#  PUT settle  → admin, manager
#  PUT update  → admin, manager
#  DELETE      → admin only
# ═════════════════════════════════════════════════════════════

@api_router.get("/daybook/stats")
async def get_daybook_stats(current_user: User = Depends(get_current_user)):
    """Summary stats — admin/manager see all, staff/billing see own."""
    query = {} if current_user.role in ['admin', 'manager'] else {'created_by': current_user.id}
    entries = await db.daybook.find(query, {'_id': 0}).to_list(10000)

    total_borrowed = sum(e['amount'] for e in entries if e['type'] == 'borrow')
    total_given    = sum(e['amount'] for e in entries if e['type'] == 'give')
    pending_amount = sum(e['amount'] for e in entries if e['type'] == 'give' and e['status'] == 'pending')
    pending_count  = sum(1 for e in entries if e['status'] == 'pending')

    today_str     = datetime.now(timezone.utc).strftime('%Y-%m-%d')
    today_borrowed = sum(e['amount'] for e in entries if e['type'] == 'borrow' and e.get('date', '') == today_str)
    today_given    = sum(e['amount'] for e in entries if e['type'] == 'give'   and e.get('date', '') == today_str)

    return {
        'total_borrowed': round(total_borrowed, 2),
        'total_given':    round(total_given, 2),
        'net_balance':    round(total_borrowed - total_given, 2),
        'pending_amount': round(pending_amount, 2),
        'pending_count':  pending_count,
        'today_borrowed': round(today_borrowed, 2),
        'today_given':    round(today_given, 2),
        'total_entries':  len(entries),
    }


@api_router.get("/daybook")
async def get_daybook_entries(current_user: User = Depends(get_current_user)):
    """Admin/manager: all entries. Staff/billing: own entries only."""
    query = {} if current_user.role in ['admin', 'manager'] else {'created_by': current_user.id}
    entries = await db.daybook.find(query, {'_id': 0}).sort('created_at', -1).to_list(1000)
    for e in entries:
        if isinstance(e.get('created_at'), str):
            e['created_at'] = datetime.fromisoformat(e['created_at'])
    return entries


@api_router.post("/daybook", response_model=DayBookEntry)
async def create_daybook_entry(
    entry_data: DayBookEntryCreate,
    current_user: User = Depends(get_current_user)
):
    require_roles(current_user, ['admin', 'manager'],
                  "Only admins and managers can create day book entries")

    if entry_data.amount <= 0:
        raise HTTPException(status_code=400, detail="Amount must be positive")
    if entry_data.type not in ('borrow', 'give'):
        raise HTTPException(status_code=400, detail="Type must be 'borrow' or 'give'")
    if entry_data.party_type not in ('staff', 'vendor', 'other'):
        raise HTTPException(status_code=400, detail="Invalid party_type")

    assigned_to_name = None
    if entry_data.assigned_to_user_id:
        staff = await db.users.find_one({'id': entry_data.assigned_to_user_id}, {'_id': 0})
        if staff:
            assigned_to_name = staff['name']

    entry = DayBookEntry(
        type=entry_data.type,
        amount=entry_data.amount,
        party_name=entry_data.party_name.strip(),
        party_type=entry_data.party_type,
        assigned_to_user_id=entry_data.assigned_to_user_id,
        assigned_to_name=assigned_to_name,
        notes=entry_data.notes,
        date=entry_data.date,
        created_by=current_user.id,
        created_by_name=current_user.name,
    )
    doc = entry.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    await db.daybook.insert_one(doc)
    return entry


@api_router.put("/daybook/{entry_id}/settle")
async def settle_daybook_entry(entry_id: str, current_user: User = Depends(get_current_user)):
    require_roles(current_user, ['admin', 'manager'],
                  "Only admins and managers can settle entries")

    entry = await db.daybook.find_one({'id': entry_id}, {'_id': 0})
    if not entry:
        raise HTTPException(status_code=404, detail="Entry not found")

    now_str = datetime.now(timezone.utc).isoformat()
    await db.daybook.update_one(
        {'id': entry_id},
        {'$set': {'status': 'settled', 'settled_at': now_str}}
    )
    return {'message': 'Entry settled successfully'}


@api_router.put("/daybook/{entry_id}")
async def update_daybook_entry(
    entry_id: str,
    update_data: DayBookEntryCreate,
    current_user: User = Depends(get_current_user)
):
    require_roles(current_user, ['admin', 'manager'],
                  "Only admins and managers can update entries")

    entry = await db.daybook.find_one({'id': entry_id}, {'_id': 0})
    if not entry:
        raise HTTPException(status_code=404, detail="Entry not found")

    update_dict = {k: v for k, v in update_data.model_dump().items() if v is not None}
    await db.daybook.update_one({'id': entry_id}, {'$set': update_dict})
    updated = await db.daybook.find_one({'id': entry_id}, {'_id': 0})
    if isinstance(updated.get('created_at'), str):
        updated['created_at'] = datetime.fromisoformat(updated['created_at'])
    return updated


@api_router.delete("/daybook/{entry_id}")
async def delete_daybook_entry(entry_id: str, current_user: User = Depends(get_current_user)):
    require_roles(current_user, ['admin'], "Only admins can delete day book entries")

    entry = await db.daybook.find_one({'id': entry_id}, {'_id': 0})
    if not entry:
        raise HTTPException(status_code=404, detail="Entry not found")

    await db.daybook.delete_one({'id': entry_id})
    return {'message': 'Entry deleted successfully'}


# ─── App setup ────────────────────────────────────────────────
app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)


@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()