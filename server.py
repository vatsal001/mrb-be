from fastapi import FastAPI, APIRouter, HTTPException, status, Depends, UploadFile, File, Response
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import logging
from pathlib import Path
from pydantic import BaseModel, Field, ConfigDict
from typing import List, Optional
from enum import Enum
import uuid
from datetime import datetime, timezone, timedelta
import bcrypt
import jwt
import barcode
from barcode.writer import ImageWriter
from io import BytesIO
import base64
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

app = FastAPI()
api_router = APIRouter(prefix="/api")
security = HTTPBearer()

JWT_SECRET = os.environ.get('JWT_SECRET', 'your-secret-key-change-in-production')

# ─────────────────────────────────────────────────────────────
#  ROLES
#  admin   → full access (all CRUD, users, reports, delete)
#  billing → POS sales + view own orders only
#  manager → inventory, stock, locations, racks, transfers, daybook, reports
#  staff   → read-only on products, inventory, daybook, locations
# ─────────────────────────────────────────────────────────────

VALID_ROLES = ["admin", "billing", "manager", "staff"]

def require_roles(user, allowed: list, detail: str = "Access denied"):
    """Raise 403 if user.role is not in allowed list."""
    if user.role not in allowed:
        raise HTTPException(status_code=403, detail=detail)

class CommissionCreate(BaseModel):
    staff_id: str
    amount: float
    commission_type: str = "flat"          # "flat" | "percentage"
    percentage_value: Optional[float] = None
    order_id: Optional[str] = None
    order_amount: Optional[float] = None
    notes: str = ""
    date: str                              # YYYY-MM-DD
 
class BulkPayRequest(BaseModel):
    commission_ids: list

class AttendanceRecord(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    user_id: str
    user_name: str
    date: str                           # YYYY-MM-DD
    clock_in: Optional[str] = None      # ISO datetime string
    clock_out: Optional[str] = None     # ISO datetime string
    duration_minutes: Optional[int] = None
    status: str = "present"             # 'present' | 'half_day' | 'on_leave'
    notes: Optional[str] = ""
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))


class LeaveRequest(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    user_id: str
    user_name: str
    date_from: str                      # YYYY-MM-DD
    date_to: str                        # YYYY-MM-DD (same as date_from for single day)
    days_count: int = 1
    leave_type: str                     # 'sick' | 'casual' | 'emergency' | 'other'
    reason: str
    status: str = "pending"             # 'pending' | 'approved' | 'rejected'
    reviewed_by_id: Optional[str] = None
    reviewed_by_name: Optional[str] = None
    reviewed_at: Optional[str] = None
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))


class LeaveRequestCreate(BaseModel):
    date_from: str
    date_to: str
    leave_type: str
    reason: str

class UserRole(str, Enum):
    ADMIN   = "admin"
    BILLING = "billing"
    MANAGER = "manager"
    STAFF   = "staff"


# ─── Models ───────────────────────────────────────────────────

class DayBookEntry(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    type: str
    amount: float
    party_name: str
    party_type: str
    assigned_to_user_id: Optional[str] = None
    assigned_to_name: Optional[str] = None
    notes: Optional[str] = ""
    date: str
    status: str = "pending"
    created_by: str
    created_by_name: str
    settled_at: Optional[str] = None
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))


class DayBookEntryCreate(BaseModel):
    type: str
    amount: float
    party_name: str
    party_type: str
    assigned_to_user_id: Optional[str] = None
    notes: Optional[str] = ""
    date: str


class User(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    email: str
    name: str
    role: str = "staff"
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))


class UserCreate(BaseModel):
    email: str
    password: str
    name: str
    role: Optional[str] = "staff"


class UserLogin(BaseModel):
    email: str
    password: str


class Product(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    name: str
    sku: str
    barcode: str
    category: str
    purchase_price: float
    selling_price: float
    stock_quantity: int
    supplier: Optional[str] = ""
    image_url: Optional[str] = ""
    low_stock_threshold: int = 10
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))


class ProductCreate(BaseModel):
    name: str
    sku: str
    category: str
    purchase_price: float
    selling_price: float
    stock_quantity: int
    supplier: Optional[str] = ""
    image_url: Optional[str] = ""
    low_stock_threshold: Optional[int] = 10


class ProductUpdate(BaseModel):
    name: Optional[str] = None
    sku: Optional[str] = None
    category: Optional[str] = None
    purchase_price: Optional[float] = None
    selling_price: Optional[float] = None
    stock_quantity: Optional[int] = None
    supplier: Optional[str] = None
    image_url: Optional[str] = None
    low_stock_threshold: Optional[int] = None


class OrderItem(BaseModel):
    product_id: str
    product_name: str
    quantity: int
    price: float
    total: float
    # GST fields (optional — POS sends these, old orders won't have them)
    hsn: Optional[str] = ""
    unit: Optional[str] = "Nos"
    discount: Optional[float] = 0
    gst_rate: Optional[float] = 18
    cgst: Optional[float] = 0
    sgst: Optional[float] = 0
    igst: Optional[float] = 0
    taxable: Optional[float] = 0


class Order(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    invoice_number: str
    items: List[OrderItem]
    subtotal: float
    tax: float
    discount: float
    total: float
    # Extended GST fields
    cgst: Optional[float] = 0
    sgst: Optional[float] = 0
    igst: Optional[float] = 0
    total_tax: Optional[float] = 0
    round_off: Optional[float] = 0
    customer_name: Optional[str] = ""
    salesman_id: Optional[str] = ""
    narration: Optional[str] = ""
    payment_mode: Optional[str] = "Cash"
    gst_type: Optional[str] = "CGST+SGST"
    invoice_date: Optional[str] = ""
    created_by: str
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))


class OrderCreate(BaseModel):
    items: List[OrderItem]
    subtotal: float
    tax: float
    discount: float
    total: float
    cgst: Optional[float] = 0
    sgst: Optional[float] = 0
    igst: Optional[float] = 0
    total_tax: Optional[float] = 0
    round_off: Optional[float] = 0
    customer_name: Optional[str] = ""
    salesman_id: Optional[str] = ""
    narration: Optional[str] = ""
    payment_mode: Optional[str] = "Cash"
    gst_type: Optional[str] = "CGST+SGST"
    invoice_date: Optional[str] = ""


class Location(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    name: str
    type: str
    description: Optional[str] = ""
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))


class LocationCreate(BaseModel):
    name: str
    type: str
    description: Optional[str] = ""


class Rack(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    code: str
    name: str
    location_id: str
    location_name: str
    description: Optional[str] = ""
    max_capacity: Optional[int] = None
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))


class RackCreate(BaseModel):
    code: str
    name: str
    location_id: str
    description: Optional[str] = ""
    max_capacity: Optional[int] = None


class RackUpdate(BaseModel):
    code: Optional[str] = None
    name: Optional[str] = None
    location_id: Optional[str] = None
    description: Optional[str] = None
    max_capacity: Optional[int] = None


class RackAssignment(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    product_id: str
    product_name: str
    rack_id: str
    rack_code: str
    location_id: str
    location_name: str
    quantity: int
    assigned_by: str
    assigned_by_name: str
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    updated_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))


class RackAssignmentCreate(BaseModel):
    product_id: str
    rack_id: str
    quantity: int


class RackAssignmentUpdate(BaseModel):
    quantity: int


class StockTransfer(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    product_id: str
    product_name: str
    from_rack_id: Optional[str] = None
    from_rack_code: Optional[str] = None
    to_rack_id: Optional[str] = None
    to_rack_code: Optional[str] = None
    quantity: int
    transfer_type: str
    notes: Optional[str] = ""
    transferred_by: str
    transferred_by_name: str
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))


class StockTransferCreate(BaseModel):
    product_id: str
    from_rack_id: Optional[str] = None
    to_rack_id: Optional[str] = None
    quantity: int
    transfer_type: str = 'manual'
    notes: Optional[str] = ""


class ProductLocationThreshold(BaseModel):
    product_id: str
    mall_threshold: int = 5
    warehouse_threshold: int = 20

# ─── Normal Helpers ─────────────────────────────────────────────

def _count_working_days(date_from: str, date_to: str) -> int:
    """Count calendar days between two dates inclusive."""
    from datetime import date as dt_date
    d1 = dt_date.fromisoformat(date_from)
    d2 = dt_date.fromisoformat(date_to)
    return max(1, (d2 - d1).days + 1)


def _today_str() -> str:
    return datetime.now(timezone.utc).strftime("%Y-%m-%d")


def _now_iso() -> str:
    return datetime.now(timezone.utc).isoformat()


# ─── Auth Helpers ─────────────────────────────────────────────

def hash_password(password: str) -> str:
    return bcrypt.hashpw(password.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')


def verify_password(password: str, hashed: str) -> bool:
    return bcrypt.checkpw(password.encode('utf-8'), hashed.encode('utf-8'))


def create_token(user_id: str, email: str, role: str) -> str:
    payload = {
        'user_id': user_id,
        'email': email,
        'role': role,
        'exp': datetime.now(timezone.utc) + timedelta(days=7)
    }
    return jwt.encode(payload, JWT_SECRET, algorithm='HS256')


async def get_current_user(credentials: HTTPAuthorizationCredentials = Depends(security)):
    try:
        token = credentials.credentials
        payload = jwt.decode(token, JWT_SECRET, algorithms=['HS256'])
        user = await db.users.find_one({'id': payload['user_id']}, {'_id': 0})
        if not user:
            raise HTTPException(status_code=401, detail="User not found")
        return User(**user)
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Token expired")
    except Exception:
        raise HTTPException(status_code=401, detail="Invalid token")


def generate_barcode_image(code: str) -> str:
    EAN = barcode.get_barcode_class('code128')
    ean = EAN(code, writer=ImageWriter())
    buffer = BytesIO()
    ean.write(buffer)
    buffer.seek(0)
    return base64.b64encode(buffer.getvalue()).decode('utf-8')


# ═════════════════════════════════════════════════════════════
#  AUTH ENDPOINTS
# ═════════════════════════════════════════════════════════════

@api_router.post("/auth/register", response_model=User)
async def register(
    user_data: UserCreate,
    current_user: Optional[User] = None   # optional — first admin is created freely
):
    existing = await db.users.find_one({'email': user_data.email}, {'_id': 0})
    if existing:
        raise HTTPException(status_code=400, detail="Email already registered")

    # Validate role
    role = user_data.role if user_data.role in VALID_ROLES else "staff"

    user = User(email=user_data.email, name=user_data.name, role=role)
    doc = user.model_dump()
    doc['password'] = hash_password(user_data.password)
    doc['created_at'] = doc['created_at'].isoformat()
    await db.users.insert_one(doc)
    return user


@api_router.post("/auth/login")
async def login(credentials: UserLogin):
    user_doc = await db.users.find_one({'email': credentials.email}, {'_id': 0})
    if not user_doc or not verify_password(credentials.password, user_doc['password']):
        raise HTTPException(status_code=401, detail="Invalid credentials")

    token = create_token(user_doc['id'], user_doc['email'], user_doc['role'])
    return {
        'token': token,
        'user': {
            'id':    user_doc['id'],
            'email': user_doc['email'],
            'name':  user_doc['name'],
            'role':  user_doc['role'],
        }
    }


@api_router.get("/auth/me", response_model=User)
async def get_me(current_user: User = Depends(get_current_user)):
    return current_user


# ═════════════════════════════════════════════════════════════
#  PRODUCTS
#  GET    → all roles
#  POST   → admin, manager
#  PUT    → admin, manager
#  DELETE → admin only
# ═════════════════════════════════════════════════════════════

@api_router.get("/products", response_model=List[Product])
async def get_products(current_user: User = Depends(get_current_user)):
    # All roles can view products
    products = await db.products.find({}, {'_id': 0}).to_list(1000)
    for p in products:
        if isinstance(p.get('created_at'), str):
            p['created_at'] = datetime.fromisoformat(p['created_at'])
    return products


@api_router.post("/products", response_model=Product)
async def create_product(
    product_data: ProductCreate,
    current_user: User = Depends(get_current_user)
):
    require_roles(current_user, ['admin', 'manager'],
                  "Only admins and managers can create products")

    barcode_num = str(uuid.uuid4().int)[:12]
    product = Product(**product_data.model_dump(), barcode=barcode_num)
    doc = product.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    await db.products.insert_one(doc)
    return product


@api_router.get("/products/{product_id}", response_model=Product)
async def get_product(product_id: str, current_user: User = Depends(get_current_user)):
    product = await db.products.find_one({'id': product_id}, {'_id': 0})
    if not product:
        raise HTTPException(status_code=404, detail="Product not found")
    if isinstance(product.get('created_at'), str):
        product['created_at'] = datetime.fromisoformat(product['created_at'])
    return Product(**product)


@api_router.get("/products/barcode/{barcode_num}")
async def get_product_by_barcode(barcode_num: str, current_user: User = Depends(get_current_user)):
    product = await db.products.find_one({'barcode': barcode_num}, {'_id': 0})
    if not product:
        raise HTTPException(status_code=404, detail="Product not found")
    if isinstance(product.get('created_at'), str):
        product['created_at'] = datetime.fromisoformat(product['created_at'])
    return product


@api_router.put("/products/{product_id}", response_model=Product)
async def update_product(
    product_id: str,
    update_data: ProductUpdate,
    current_user: User = Depends(get_current_user)
):
    require_roles(current_user, ['admin', 'manager'],
                  "Only admins and managers can update products")

    existing = await db.products.find_one({'id': product_id}, {'_id': 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Product not found")

    update_dict = {k: v for k, v in update_data.model_dump().items() if v is not None}
    if update_dict:
        await db.products.update_one({'id': product_id}, {'$set': update_dict})

    updated = await db.products.find_one({'id': product_id}, {'_id': 0})
    if isinstance(updated.get('created_at'), str):
        updated['created_at'] = datetime.fromisoformat(updated['created_at'])
    return Product(**updated)


@api_router.delete("/products/{product_id}")
async def delete_product(product_id: str, current_user: User = Depends(get_current_user)):
    require_roles(current_user, ['admin'], "Only admins can delete products")

    result = await db.products.delete_one({'id': product_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Product not found")
    return {'message': 'Product deleted successfully'}


@api_router.get("/products/{product_id}/barcode-image")
async def get_barcode_image(product_id: str, current_user: User = Depends(get_current_user)):
    product = await db.products.find_one({'id': product_id}, {'_id': 0})
    if not product:
        raise HTTPException(status_code=404, detail="Product not found")
    barcode_img = generate_barcode_image(product['barcode'])
    return {'barcode_image': f'data:image/png;base64,{barcode_img}'}


# ═════════════════════════════════════════════════════════════
#  ORDERS / POS
#  POST (checkout) → admin, billing
#  GET (list)      → admin, billing (own only), manager (view all)
# ═════════════════════════════════════════════════════════════

@api_router.post("/orders", response_model=Order)
async def create_order(
    order_data: OrderCreate,
    current_user: User = Depends(get_current_user)
):
    require_roles(current_user, ['admin', 'billing'],
                  "Only admins and billing staff can create orders")

    invoice_num = f"INV-{datetime.now().strftime('%Y%m%d')}-{str(uuid.uuid4())[:8].upper()}"

    order = Order(
        invoice_number=invoice_num,
        items=order_data.items,
        subtotal=order_data.subtotal,
        tax=order_data.tax,
        discount=order_data.discount,
        total=order_data.total,
        cgst=order_data.cgst,
        sgst=order_data.sgst,
        igst=order_data.igst,
        total_tax=order_data.total_tax,
        round_off=order_data.round_off,
        customer_name=order_data.customer_name,
        salesman_id=order_data.salesman_id,
        narration=order_data.narration,
        payment_mode=order_data.payment_mode,
        gst_type=order_data.gst_type,
        invoice_date=order_data.invoice_date,
        created_by=current_user.id
    )

    # Deduct stock from mall rack assignments
    mall_locations = await db.locations.find({'type': 'mall'}, {'_id': 0}).to_list(10)
    mall_location_ids = [loc['id'] for loc in mall_locations]

    for item in order_data.items:
        product = await db.products.find_one({'id': item.product_id}, {'_id': 0})
        if product:
            new_stock = product['stock_quantity'] - item.quantity
            await db.products.update_one(
                {'id': item.product_id},
                {'$set': {'stock_quantity': max(0, new_stock)}}
            )

            remaining_qty = item.quantity
            mall_assignments = await db.rack_assignments.find({
                'product_id': item.product_id,
                'location_id': {'$in': mall_location_ids}
            }, {'_id': 0}).sort('quantity', -1).to_list(100)

            for assignment in mall_assignments:
                if remaining_qty <= 0:
                    break
                deduct_qty = min(remaining_qty, assignment['quantity'])
                new_rack_qty = assignment['quantity'] - deduct_qty
                if new_rack_qty == 0:
                    await db.rack_assignments.delete_one({'id': assignment['id']})
                else:
                    await db.rack_assignments.update_one(
                        {'id': assignment['id']},
                        {'$set': {
                            'quantity': new_rack_qty,
                            'updated_at': datetime.now(timezone.utc).isoformat()
                        }}
                    )

                transfer = StockTransfer(
                    product_id=item.product_id,
                    product_name=product['name'],
                    from_rack_id=assignment['rack_id'],
                    from_rack_code=assignment['rack_code'],
                    quantity=deduct_qty,
                    transfer_type='sale',
                    notes=f"Sale deduction from invoice {invoice_num}",
                    transferred_by=current_user.id,
                    transferred_by_name=current_user.name
                )
                transfer_doc = transfer.model_dump()
                transfer_doc['created_at'] = transfer_doc['created_at'].isoformat()
                await db.stock_transfers.insert_one(transfer_doc)
                remaining_qty -= deduct_qty

    doc = order.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    await db.orders.insert_one(doc)
    return order


@api_router.get("/orders", response_model=List[Order])
async def get_orders(current_user: User = Depends(get_current_user)):
    require_roles(current_user, ['admin', 'billing', 'manager'],
                  "Staff cannot view orders")

    # billing staff see only their own orders
    query = {} if current_user.role in ['admin', 'manager'] else {'created_by': current_user.id}
    orders = await db.orders.find(query, {'_id': 0}).sort('created_at', -1).to_list(1000)
    for order in orders:
        if isinstance(order.get('created_at'), str):
            order['created_at'] = datetime.fromisoformat(order['created_at'])
    return orders


@api_router.get("/orders/{order_id}", response_model=Order)
async def get_order(order_id: str, current_user: User = Depends(get_current_user)):
    require_roles(current_user, ['admin', 'billing', 'manager'],
                  "Staff cannot view orders")

    order = await db.orders.find_one({'id': order_id}, {'_id': 0})
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")
    # billing can only see their own orders
    if current_user.role == 'billing' and order.get('created_by') != current_user.id:
        raise HTTPException(status_code=403, detail="Access denied")
    if isinstance(order.get('created_at'), str):
        order['created_at'] = datetime.fromisoformat(order['created_at'])
    return Order(**order)


# ═════════════════════════════════════════════════════════════
#  REPORTS
#  admin, manager only
# ═════════════════════════════════════════════════════════════

@api_router.get("/reports/sales")
async def get_sales_report(
    period: str = "daily",
    current_user: User = Depends(get_current_user)
):
    require_roles(current_user, ['admin', 'manager'],
                  "Only admins and managers can view reports")

    now = datetime.now(timezone.utc)
    if period == "daily":
        start_date = now.replace(hour=0, minute=0, second=0, microsecond=0)
    elif period == "weekly":
        start_date = now - timedelta(days=7)
    elif period == "monthly":
        start_date = now - timedelta(days=30)
    else:
        start_date = now - timedelta(days=30)

    orders = await db.orders.find({}, {'_id': 0}).to_list(10000)
    filtered_orders = []
    for order in orders:
        created_at = order.get('created_at')
        if isinstance(created_at, str):
            created_at = datetime.fromisoformat(created_at)
        if created_at >= start_date:
            filtered_orders.append(order)

    total_sales = sum(order['total'] for order in filtered_orders)
    total_orders = len(filtered_orders)

    products = await db.products.find({}, {'_id': 0}).to_list(1000)
    total_profit = 0
    for order in filtered_orders:
        for item in order['items']:
            product = next((p for p in products if p['id'] == item['product_id']), None)
            if product:
                profit = (item['price'] - product['purchase_price']) * item['quantity']
                total_profit += profit

    return {
        'period': period,
        'total_sales': round(total_sales, 2),
        'total_orders': total_orders,
        'total_profit': round(total_profit, 2),
        'orders': filtered_orders
    }


@api_router.get("/reports/export/excel")
async def export_excel_report(
    period: str = "daily",
    current_user: User = Depends(get_current_user)
):
    require_roles(current_user, ['admin', 'manager'],
                  "Only admins and managers can export reports")

    report_data = await get_sales_report(period, current_user)

    wb = Workbook()
    ws = wb.active
    ws.title = f"{period.capitalize()} Sales Report"
    ws['A1'] = f"Sales Report - {period.capitalize()}"
    ws['A1'].font = Font(size=14, bold=True)
    ws['A3'] = 'Total Sales:'
    ws['B3'] = report_data['total_sales']
    ws['A4'] = 'Total Orders:'
    ws['B4'] = report_data['total_orders']
    ws['A5'] = 'Total Profit:'
    ws['B5'] = report_data['total_profit']

    headers = ['Invoice Number', 'Date', 'Items', 'Subtotal', 'Tax', 'Discount', 'Total']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=7, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    for row_idx, order in enumerate(report_data['orders'], start=8):
        ws.cell(row=row_idx, column=1, value=order['invoice_number'])
        created_at = order['created_at']
        if isinstance(created_at, str):
            created_at = datetime.fromisoformat(created_at)
        ws.cell(row=row_idx, column=2, value=created_at.strftime('%Y-%m-%d %H:%M'))
        ws.cell(row=row_idx, column=3, value=len(order['items']))
        ws.cell(row=row_idx, column=4, value=order['subtotal'])
        ws.cell(row=row_idx, column=5, value=order['tax'])
        ws.cell(row=row_idx, column=6, value=order['discount'])
        ws.cell(row=row_idx, column=7, value=order['total'])

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return Response(
        content=buffer.getvalue(),
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename=sales_report_{period}.xlsx'}
    )


# ═════════════════════════════════════════════════════════════
#  DASHBOARD
#  All roles (financial totals hidden for billing/staff in frontend)
# ═════════════════════════════════════════════════════════════

@api_router.get("/dashboard/stats")
async def get_dashboard_stats(current_user: User = Depends(get_current_user)):
    products = await db.products.find({}, {'_id': 0}).to_list(1000)
    orders   = await db.orders.find({}, {'_id': 0}).to_list(1000)

    now         = datetime.now(timezone.utc)
    today_start = now.replace(hour=0, minute=0, second=0, microsecond=0)

    today_orders = []
    for order in orders:
        created_at = order.get('created_at')
        if isinstance(created_at, str):
            created_at = datetime.fromisoformat(created_at)
        if created_at >= today_start:
            today_orders.append(order)

    today_sales = sum(o['total'] for o in today_orders)
    total_sales = sum(o['total'] for o in orders)

    total_profit = 0
    for order in orders:
        for item in order['items']:
            product = next((p for p in products if p['id'] == item['product_id']), None)
            if product:
                profit = (item['price'] - product['purchase_price']) * item['quantity']
                total_profit += profit

    low_stock_products = [p for p in products if p['stock_quantity'] <= p.get('low_stock_threshold', 10)]

    return {
        'today_sales':        round(today_sales, 2),
        'total_sales':        round(total_sales, 2),
        'total_profit':       round(total_profit, 2),
        'total_products':     len(products),
        'total_orders':       len(orders),
        'low_stock_count':    len(low_stock_products),
        'low_stock_products': low_stock_products[:5],
        # expose role so frontend can conditionally hide financial data
        'viewer_role':        current_user.role,
    }


# ═════════════════════════════════════════════════════════════
#  USERS
#  GET, POST, PUT role → admin only
# ═════════════════════════════════════════════════════════════

@api_router.get("/users", response_model=List[User])
async def get_users(current_user: User = Depends(get_current_user)):
    require_roles(current_user, ['admin'], "Only admins can view users")

    users = await db.users.find({}, {'_id': 0, 'password': 0}).to_list(1000)
    for u in users:
        if isinstance(u.get('created_at'), str):
            u['created_at'] = datetime.fromisoformat(u['created_at'])
    return users


@api_router.put("/users/{user_id}/role")
async def update_user_role(
    user_id: str,
    role: str,
    current_user: User = Depends(get_current_user)
):
    require_roles(current_user, ['admin'], "Only admins can update roles")

    if role not in VALID_ROLES:
        raise HTTPException(
            status_code=400,
            detail=f"Invalid role '{role}'. Must be one of: {', '.join(VALID_ROLES)}"
        )

    result = await db.users.update_one({'id': user_id}, {'$set': {'role': role}})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="User not found")
    return {'message': f'Role updated to {role}'}


# ═════════════════════════════════════════════════════════════
#  LOCATIONS
#  GET  → all roles
#  POST → admin, manager
# ═════════════════════════════════════════════════════════════

@api_router.get("/locations", response_model=List[Location])
async def get_locations(current_user: User = Depends(get_current_user)):
    locations = await db.locations.find({}, {'_id': 0}).to_list(1000)
    for loc in locations:
        if isinstance(loc.get('created_at'), str):
            loc['created_at'] = datetime.fromisoformat(loc['created_at'])
    return locations


@api_router.post("/locations", response_model=Location)
async def create_location(
    location_data: LocationCreate,
    current_user: User = Depends(get_current_user)
):
    require_roles(current_user, ['admin', 'manager'],
                  "Only admins and managers can create locations")

    location = Location(**location_data.model_dump())
    doc = location.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    await db.locations.insert_one(doc)
    return location


# ═════════════════════════════════════════════════════════════
#  RACKS
#  GET    → all roles
#  POST   → admin, manager
#  PUT    → admin, manager
#  DELETE → admin only
# ═════════════════════════════════════════════════════════════

@api_router.get("/racks", response_model=List[Rack])
async def get_racks(
    location_id: Optional[str] = None,
    current_user: User = Depends(get_current_user)
):
    query = {'location_id': location_id} if location_id else {}
    racks = await db.racks.find(query, {'_id': 0}).to_list(1000)
    for rack in racks:
        if isinstance(rack.get('created_at'), str):
            rack['created_at'] = datetime.fromisoformat(rack['created_at'])
    return racks


@api_router.post("/racks", response_model=Rack)
async def create_rack(
    rack_data: RackCreate,
    current_user: User = Depends(get_current_user)
):
    require_roles(current_user, ['admin', 'manager'],
                  "Only admins and managers can create racks")

    location = await db.locations.find_one({'id': rack_data.location_id}, {'_id': 0})
    if not location:
        raise HTTPException(status_code=404, detail="Location not found")

    existing = await db.racks.find_one({
        'code': rack_data.code, 'location_id': rack_data.location_id
    }, {'_id': 0})
    if existing:
        raise HTTPException(status_code=400, detail="Rack code already exists in this location")

    rack = Rack(**rack_data.model_dump(), location_name=location['name'])
    doc = rack.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    await db.racks.insert_one(doc)
    return rack


@api_router.get("/racks/{rack_id}", response_model=Rack)
async def get_rack(rack_id: str, current_user: User = Depends(get_current_user)):
    rack = await db.racks.find_one({'id': rack_id}, {'_id': 0})
    if not rack:
        raise HTTPException(status_code=404, detail="Rack not found")
    if isinstance(rack.get('created_at'), str):
        rack['created_at'] = datetime.fromisoformat(rack['created_at'])
    return Rack(**rack)


@api_router.put("/racks/{rack_id}", response_model=Rack)
async def update_rack(
    rack_id: str,
    update_data: RackUpdate,
    current_user: User = Depends(get_current_user)
):
    require_roles(current_user, ['admin', 'manager'],
                  "Only admins and managers can update racks")

    existing = await db.racks.find_one({'id': rack_id}, {'_id': 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Rack not found")

    update_dict = {k: v for k, v in update_data.model_dump().items() if v is not None}
    if 'location_id' in update_dict:
        location = await db.locations.find_one({'id': update_dict['location_id']}, {'_id': 0})
        if location:
            update_dict['location_name'] = location['name']
    if update_dict:
        await db.racks.update_one({'id': rack_id}, {'$set': update_dict})

    updated = await db.racks.find_one({'id': rack_id}, {'_id': 0})
    if isinstance(updated.get('created_at'), str):
        updated['created_at'] = datetime.fromisoformat(updated['created_at'])
    return Rack(**updated)


@api_router.delete("/racks/{rack_id}")
async def delete_rack(rack_id: str, current_user: User = Depends(get_current_user)):
    require_roles(current_user, ['admin'], "Only admins can delete racks")

    assignments = await db.rack_assignments.find_one({'rack_id': rack_id}, {'_id': 0})
    if assignments:
        raise HTTPException(status_code=400, detail="Cannot delete rack with product assignments")

    result = await db.racks.delete_one({'id': rack_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Rack not found")
    return {'message': 'Rack deleted successfully'}


# ═════════════════════════════════════════════════════════════
#  RACK ASSIGNMENTS
#  GET    → all roles
#  POST   → admin, manager
#  PUT    → admin, manager
#  DELETE → admin, manager
# ═════════════════════════════════════════════════════════════

@api_router.get("/products/{product_id}/rack-assignments", response_model=List[RackAssignment])
async def get_product_rack_assignments(product_id: str, current_user: User = Depends(get_current_user)):
    assignments = await db.rack_assignments.find({'product_id': product_id}, {'_id': 0}).to_list(1000)
    for a in assignments:
        if isinstance(a.get('created_at'), str):
            a['created_at'] = datetime.fromisoformat(a['created_at'])
        if isinstance(a.get('updated_at'), str):
            a['updated_at'] = datetime.fromisoformat(a['updated_at'])
    return assignments


@api_router.get("/racks/{rack_id}/products")
async def get_rack_products(rack_id: str, current_user: User = Depends(get_current_user)):
    return await db.rack_assignments.find({'rack_id': rack_id}, {'_id': 0}).to_list(1000)


@api_router.post("/rack-assignments", response_model=RackAssignment)
async def create_rack_assignment(
    assignment_data: RackAssignmentCreate,
    current_user: User = Depends(get_current_user)
):
    require_roles(current_user, ['admin', 'manager'],
                  "Only admins and managers can assign products to racks")

    product = await db.products.find_one({'id': assignment_data.product_id}, {'_id': 0})
    if not product:
        raise HTTPException(status_code=404, detail="Product not found")
    rack = await db.racks.find_one({'id': assignment_data.rack_id}, {'_id': 0})
    if not rack:
        raise HTTPException(status_code=404, detail="Rack not found")

    existing = await db.rack_assignments.find_one({
        'product_id': assignment_data.product_id,
        'rack_id':    assignment_data.rack_id
    }, {'_id': 0})
    if existing:
        raise HTTPException(status_code=400, detail="Product already assigned to this rack")

    assignments = await db.rack_assignments.find({'product_id': assignment_data.product_id}, {'_id': 0}).to_list(1000)
    total_assigned = sum(a['quantity'] for a in assignments) + assignment_data.quantity
    if total_assigned > product['stock_quantity']:
        raise HTTPException(
            status_code=400,
            detail=f"Total assigned ({total_assigned}) exceeds available stock ({product['stock_quantity']})"
        )

    assignment = RackAssignment(
        product_id=assignment_data.product_id,
        product_name=product['name'],
        rack_id=assignment_data.rack_id,
        rack_code=rack['code'],
        location_id=rack['location_id'],
        location_name=rack['location_name'],
        quantity=assignment_data.quantity,
        assigned_by=current_user.id,
        assigned_by_name=current_user.name
    )
    doc = assignment.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    doc['updated_at'] = doc['updated_at'].isoformat()
    await db.rack_assignments.insert_one(doc)
    return assignment


@api_router.put("/rack-assignments/{assignment_id}", response_model=RackAssignment)
async def update_rack_assignment(
    assignment_id: str,
    update_data: RackAssignmentUpdate,
    current_user: User = Depends(get_current_user)
):
    require_roles(current_user, ['admin', 'manager'],
                  "Only admins and managers can update rack assignments")

    existing = await db.rack_assignments.find_one({'id': assignment_id}, {'_id': 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Assignment not found")

    product = await db.products.find_one({'id': existing['product_id']}, {'_id': 0})
    other_assignments = await db.rack_assignments.find({
        'product_id': existing['product_id'], 'id': {'$ne': assignment_id}
    }, {'_id': 0}).to_list(1000)
    total_assigned = sum(a['quantity'] for a in other_assignments) + update_data.quantity
    if total_assigned > product['stock_quantity']:
        raise HTTPException(
            status_code=400,
            detail=f"Total assigned ({total_assigned}) exceeds available stock ({product['stock_quantity']})"
        )

    await db.rack_assignments.update_one(
        {'id': assignment_id},
        {'$set': {'quantity': update_data.quantity, 'updated_at': datetime.now(timezone.utc).isoformat()}}
    )
    updated = await db.rack_assignments.find_one({'id': assignment_id}, {'_id': 0})
    if isinstance(updated.get('created_at'), str):
        updated['created_at'] = datetime.fromisoformat(updated['created_at'])
    if isinstance(updated.get('updated_at'), str):
        updated['updated_at'] = datetime.fromisoformat(updated['updated_at'])
    return RackAssignment(**updated)


@api_router.delete("/rack-assignments/{assignment_id}")
async def delete_rack_assignment(assignment_id: str, current_user: User = Depends(get_current_user)):
    require_roles(current_user, ['admin', 'manager'],
                  "Only admins and managers can remove rack assignments")

    result = await db.rack_assignments.delete_one({'id': assignment_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Assignment not found")
    return {'message': 'Assignment deleted successfully'}


# ═════════════════════════════════════════════════════════════
#  STOCK TRANSFERS
#  GET  → all roles
#  POST → admin, manager
# ═════════════════════════════════════════════════════════════

@api_router.get("/stock-transfers", response_model=List[StockTransfer])
async def get_stock_transfers(
    product_id:  Optional[str] = None,
    start_date:  Optional[str] = None,
    end_date:    Optional[str] = None,
    current_user: User = Depends(get_current_user)
):
    query = {}
    if product_id:
        query['product_id'] = product_id

    transfers = await db.stock_transfers.find(query, {'_id': 0}).sort('created_at', -1).to_list(1000)

    if start_date or end_date:
        filtered = []
        for t in transfers:
            created_at = t.get('created_at')
            if isinstance(created_at, str):
                created_at = datetime.fromisoformat(created_at)
            if start_date and created_at < datetime.fromisoformat(start_date):
                continue
            if end_date and created_at > datetime.fromisoformat(end_date):
                continue
            filtered.append(t)
        transfers = filtered

    for t in transfers:
        if isinstance(t.get('created_at'), str):
            t['created_at'] = datetime.fromisoformat(t['created_at'])
    return transfers


@api_router.post("/stock-transfers", response_model=StockTransfer)
async def create_stock_transfer(
    transfer_data: StockTransferCreate,
    current_user: User = Depends(get_current_user)
):
    require_roles(current_user, ['admin', 'manager'],
                  "Only admins and managers can create stock transfers")

    product = await db.products.find_one({'id': transfer_data.product_id}, {'_id': 0})
    if not product:
        raise HTTPException(status_code=404, detail="Product not found")

    from_rack_code = None
    to_rack_code   = None

    if transfer_data.from_rack_id:
        from_assignment = await db.rack_assignments.find_one({
            'product_id': transfer_data.product_id, 'rack_id': transfer_data.from_rack_id
        }, {'_id': 0})
        if not from_assignment:
            raise HTTPException(status_code=404, detail="Product not found in source rack")
        if from_assignment['quantity'] < transfer_data.quantity:
            raise HTTPException(
                status_code=400,
                detail=f"Insufficient quantity. Available: {from_assignment['quantity']}"
            )
        new_quantity = from_assignment['quantity'] - transfer_data.quantity
        if new_quantity == 0:
            await db.rack_assignments.delete_one({'id': from_assignment['id']})
        else:
            await db.rack_assignments.update_one(
                {'id': from_assignment['id']},
                {'$set': {'quantity': new_quantity, 'updated_at': datetime.now(timezone.utc).isoformat()}}
            )
        from_rack = await db.racks.find_one({'id': transfer_data.from_rack_id}, {'_id': 0})
        from_rack_code = from_rack['code'] if from_rack else None

    if transfer_data.to_rack_id:
        to_assignment = await db.rack_assignments.find_one({
            'product_id': transfer_data.product_id, 'rack_id': transfer_data.to_rack_id
        }, {'_id': 0})
        to_rack = await db.racks.find_one({'id': transfer_data.to_rack_id}, {'_id': 0})
        to_rack_code = to_rack['code'] if to_rack else None

        if to_assignment:
            await db.rack_assignments.update_one(
                {'id': to_assignment['id']},
                {'$set': {'quantity': to_assignment['quantity'] + transfer_data.quantity,
                          'updated_at': datetime.now(timezone.utc).isoformat()}}
            )
        elif to_rack:
            assignment = RackAssignment(
                product_id=transfer_data.product_id,
                product_name=product['name'],
                rack_id=transfer_data.to_rack_id,
                rack_code=to_rack['code'],
                location_id=to_rack['location_id'],
                location_name=to_rack['location_name'],
                quantity=transfer_data.quantity,
                assigned_by=current_user.id,
                assigned_by_name=current_user.name
            )
            doc = assignment.model_dump()
            doc['created_at'] = doc['created_at'].isoformat()
            doc['updated_at'] = doc['updated_at'].isoformat()
            await db.rack_assignments.insert_one(doc)

    transfer = StockTransfer(
        product_id=transfer_data.product_id,
        product_name=product['name'],
        from_rack_id=transfer_data.from_rack_id,
        from_rack_code=from_rack_code,
        to_rack_id=transfer_data.to_rack_id,
        to_rack_code=to_rack_code,
        quantity=transfer_data.quantity,
        transfer_type=transfer_data.transfer_type,
        notes=transfer_data.notes,
        transferred_by=current_user.id,
        transferred_by_name=current_user.name
    )
    doc = transfer.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    await db.stock_transfers.insert_one(doc)
    return transfer


@api_router.get("/products/{product_id}/locations")
async def get_product_locations(product_id: str, current_user: User = Depends(get_current_user)):
    product = await db.products.find_one({'id': product_id}, {'_id': 0})
    if not product:
        raise HTTPException(status_code=404, detail="Product not found")

    assignments = await db.rack_assignments.find({'product_id': product_id}, {'_id': 0}).to_list(1000)
    mall_racks, warehouse_racks = [], []
    for a in assignments:
        location = await db.locations.find_one({'id': a['location_id']}, {'_id': 0})
        if location:
            if location['type'] == 'mall':
                mall_racks.append(a)
            else:
                warehouse_racks.append(a)

    mall_total      = sum(r['quantity'] for r in mall_racks)
    warehouse_total = sum(r['quantity'] for r in warehouse_racks)
    return {
        'product': product,
        'mall_racks': mall_racks,
        'warehouse_racks': warehouse_racks,
        'mall_total': mall_total,
        'warehouse_total': warehouse_total,
        'total_assigned': mall_total + warehouse_total
    }


@api_router.get("/dashboard/low-stock-by-location")
async def get_low_stock_by_location(current_user: User = Depends(get_current_user)):
    products  = await db.products.find({}, {'_id': 0}).to_list(1000)
    locations = await db.locations.find({}, {'_id': 0}).to_list(10)

    mall_low_stock, warehouse_low_stock = [], []
    for product in products:
        assignments = await db.rack_assignments.find({'product_id': product['id']}, {'_id': 0}).to_list(1000)
        mall_qty = warehouse_qty = 0
        for a in assignments:
            location = next((l for l in locations if l['id'] == a['location_id']), None)
            if location:
                if location['type'] == 'mall':
                    mall_qty += a['quantity']
                else:
                    warehouse_qty += a['quantity']

        if 0 < mall_qty <= 5:
            mall_low_stock.append({**product, 'current_quantity': mall_qty, 'threshold': 5})
        if 0 < warehouse_qty <= 20:
            warehouse_low_stock.append({**product, 'current_quantity': warehouse_qty, 'threshold': 20})

    return {
        'mall_low_stock':      mall_low_stock[:10],
        'warehouse_low_stock': warehouse_low_stock[:10],
        'mall_count':          len(mall_low_stock),
        'warehouse_count':     len(warehouse_low_stock)
    }


# ═════════════════════════════════════════════════════════════
#  DAY BOOK
#  GET         → admin, manager, staff (staff: own only)
#  POST        → admin, manager
#  PUT settle  → admin, manager
#  PUT update  → admin, manager
#  DELETE      → admin only
# ═════════════════════════════════════════════════════════════

@api_router.get("/daybook/stats")
async def get_daybook_stats(current_user: User = Depends(get_current_user)):
    """Summary stats — admin/manager see all, staff/billing see own."""
    query = {} if current_user.role in ['admin', 'manager'] else {'created_by': current_user.id}
    entries = await db.daybook.find(query, {'_id': 0}).to_list(10000)

    total_borrowed = sum(e['amount'] for e in entries if e['type'] == 'borrow')
    total_given    = sum(e['amount'] for e in entries if e['type'] == 'give')
    pending_amount = sum(e['amount'] for e in entries if e['type'] == 'give' and e['status'] == 'pending')
    pending_count  = sum(1 for e in entries if e['status'] == 'pending')

    today_str     = datetime.now(timezone.utc).strftime('%Y-%m-%d')
    today_borrowed = sum(e['amount'] for e in entries if e['type'] == 'borrow' and e.get('date', '') == today_str)
    today_given    = sum(e['amount'] for e in entries if e['type'] == 'give'   and e.get('date', '') == today_str)

    return {
        'total_borrowed': round(total_borrowed, 2),
        'total_given':    round(total_given, 2),
        'net_balance':    round(total_borrowed - total_given, 2),
        'pending_amount': round(pending_amount, 2),
        'pending_count':  pending_count,
        'today_borrowed': round(today_borrowed, 2),
        'today_given':    round(today_given, 2),
        'total_entries':  len(entries),
    }


@api_router.get("/daybook")
async def get_daybook_entries(current_user: User = Depends(get_current_user)):
    """Admin/manager: all entries. Staff/billing: own entries only."""
    query = {} if current_user.role in ['admin', 'manager'] else {'created_by': current_user.id}
    entries = await db.daybook.find(query, {'_id': 0}).sort('created_at', -1).to_list(1000)
    for e in entries:
        if isinstance(e.get('created_at'), str):
            e['created_at'] = datetime.fromisoformat(e['created_at'])
    return entries


@api_router.post("/daybook", response_model=DayBookEntry)
async def create_daybook_entry(
    entry_data: DayBookEntryCreate,
    current_user: User = Depends(get_current_user)
):
    require_roles(current_user, ['admin', 'manager'],
                  "Only admins and managers can create day book entries")

    if entry_data.amount <= 0:
        raise HTTPException(status_code=400, detail="Amount must be positive")
    if entry_data.type not in ('borrow', 'give'):
        raise HTTPException(status_code=400, detail="Type must be 'borrow' or 'give'")
    if entry_data.party_type not in ('staff', 'vendor', 'other'):
        raise HTTPException(status_code=400, detail="Invalid party_type")

    assigned_to_name = None
    if entry_data.assigned_to_user_id:
        staff = await db.users.find_one({'id': entry_data.assigned_to_user_id}, {'_id': 0})
        if staff:
            assigned_to_name = staff['name']

    entry = DayBookEntry(
        type=entry_data.type,
        amount=entry_data.amount,
        party_name=entry_data.party_name.strip(),
        party_type=entry_data.party_type,
        assigned_to_user_id=entry_data.assigned_to_user_id,
        assigned_to_name=assigned_to_name,
        notes=entry_data.notes,
        date=entry_data.date,
        created_by=current_user.id,
        created_by_name=current_user.name,
    )
    doc = entry.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    await db.daybook.insert_one(doc)
    return entry


@api_router.put("/daybook/{entry_id}/settle")
async def settle_daybook_entry(entry_id: str, current_user: User = Depends(get_current_user)):
    require_roles(current_user, ['admin', 'manager'],
                  "Only admins and managers can settle entries")

    entry = await db.daybook.find_one({'id': entry_id}, {'_id': 0})
    if not entry:
        raise HTTPException(status_code=404, detail="Entry not found")

    now_str = datetime.now(timezone.utc).isoformat()
    await db.daybook.update_one(
        {'id': entry_id},
        {'$set': {'status': 'settled', 'settled_at': now_str}}
    )
    return {'message': 'Entry settled successfully'}


@api_router.put("/daybook/{entry_id}")
async def update_daybook_entry(
    entry_id: str,
    update_data: DayBookEntryCreate,
    current_user: User = Depends(get_current_user)
):
    require_roles(current_user, ['admin', 'manager'],
                  "Only admins and managers can update entries")

    entry = await db.daybook.find_one({'id': entry_id}, {'_id': 0})
    if not entry:
        raise HTTPException(status_code=404, detail="Entry not found")

    update_dict = {k: v for k, v in update_data.model_dump().items() if v is not None}
    await db.daybook.update_one({'id': entry_id}, {'$set': update_dict})
    updated = await db.daybook.find_one({'id': entry_id}, {'_id': 0})
    if isinstance(updated.get('created_at'), str):
        updated['created_at'] = datetime.fromisoformat(updated['created_at'])
    return updated


@api_router.delete("/daybook/{entry_id}")
async def delete_daybook_entry(entry_id: str, current_user: User = Depends(get_current_user)):
    require_roles(current_user, ['admin'], "Only admins can delete day book entries")

    entry = await db.daybook.find_one({'id': entry_id}, {'_id': 0})
    if not entry:
        raise HTTPException(status_code=404, detail="Entry not found")

    await db.daybook.delete_one({'id': entry_id})
    return {'message': 'Entry deleted successfully'}


# ══════════════════════════════════════════════════════════════
#  ATTENDANCE ENDPOINTS
# ══════════════════════════════════════════════════════════════

@api_router.post("/attendance/clock-in")
async def clock_in(current_user: User = Depends(get_current_user)):
    """Mark clock-in for today. Only one clock-in per day per user."""
    today = _today_str()

    existing = await db.attendance.find_one(
        {"user_id": current_user.id, "date": today}, {"_id": 0}
    )
    if existing:
        if existing.get("clock_in"):
            raise HTTPException(status_code=400, detail="Already clocked in today")
    
    now = _now_iso()
    record = AttendanceRecord(
        user_id=current_user.id,
        user_name=current_user.name,
        date=today,
        clock_in=now,
        status="present",
    )
    doc = record.model_dump()
    doc["created_at"] = doc["created_at"].isoformat()
    await db.attendance.insert_one(doc)
    return {"message": "Clocked in successfully", "clock_in": now, "record_id": record.id}


@api_router.post("/attendance/clock-out")
async def clock_out(current_user: User = Depends(get_current_user)):
    """Mark clock-out for today."""
    today = _today_str()

    record = await db.attendance.find_one(
        {"user_id": current_user.id, "date": today}, {"_id": 0}
    )
    if not record:
        raise HTTPException(status_code=400, detail="You haven't clocked in today")
    if record.get("clock_out"):
        raise HTTPException(status_code=400, detail="Already clocked out today")

    clock_in_dt = datetime.fromisoformat(record["clock_in"])
    now_dt = datetime.now(timezone.utc)
    duration = int((now_dt - clock_in_dt).total_seconds() / 60)

    now = now_dt.isoformat()
    await db.attendance.update_one(
        {"user_id": current_user.id, "date": today},
        {"$set": {"clock_out": now, "duration_minutes": duration}}
    )
    return {"message": "Clocked out successfully", "clock_out": now, "duration_minutes": duration}


@api_router.get("/attendance/today")
async def get_today_attendance(current_user: User = Depends(get_current_user)):
    """Get today's attendance record for the current user."""
    today = _today_str()
    record = await db.attendance.find_one(
        {"user_id": current_user.id, "date": today}, {"_id": 0}
    )
    # Also check if they have an approved leave for today
    leave = await db.leaves.find_one(
        {"user_id": current_user.id, "status": "approved",
         "date_from": {"$lte": today}, "date_to": {"$gte": today}},
        {"_id": 0}
    )
    return {
        "record": record,
        "on_approved_leave": leave is not None,
        "leave": leave,
        "today": today,
    }


@api_router.get("/attendance/my")
async def get_my_attendance(
    month: Optional[str] = None,   # format: YYYY-MM
    current_user: User = Depends(get_current_user)
):
    """Get current user's attendance records, optionally filtered by month."""
    query = {"user_id": current_user.id}
    if month:
        query["date"] = {"$regex": f"^{month}"}

    records = await db.attendance.find(query, {"_id": 0}).sort("date", -1).to_list(500)
    for r in records:
        if isinstance(r.get("created_at"), str):
            r["created_at"] = datetime.fromisoformat(r["created_at"])
    return records


@api_router.get("/attendance/summary")
async def get_attendance_summary(
    month: str,                     # format: YYYY-MM  e.g. 2026-03
    current_user: User = Depends(get_current_user)
):
    """
    Admin/Manager only.
    Returns per-staff attendance summary for the given month.
    """
    require_roles(current_user, ["admin", "manager"],
                  "Only admins and managers can view attendance summary")

    # Get all users
    all_users = await db.users.find({}, {"_id": 0, "password": 0}).to_list(500)

    # Get all attendance records for the month
    records = await db.attendance.find(
        {"date": {"$regex": f"^{month}"}}, {"_id": 0}
    ).to_list(5000)

    # Get all approved leaves for the month
    leaves = await db.leaves.find(
        {"status": "approved"}, {"_id": 0}
    ).to_list(5000)

    # Build per-user summary
    summary = []
    for user in all_users:
        user_records = [r for r in records if r["user_id"] == user["id"]]

        days_present   = len([r for r in user_records if r.get("clock_in") and r.get("clock_out")])
        days_clocked_in_only = len([r for r in user_records if r.get("clock_in") and not r.get("clock_out")])

        # Count approved leave days in this month
        leave_days = 0
        user_leaves = [l for l in leaves if l["user_id"] == user["id"]]
        for leave in user_leaves:
            # Count overlap with requested month
            from datetime import date as dt_date
            try:
                d1 = max(dt_date.fromisoformat(leave["date_from"]), dt_date.fromisoformat(f"{month}-01"))
                last_day = dt_date.fromisoformat(f"{month}-01").replace(day=28) + __import__('datetime').timedelta(days=4)
                month_end = last_day - __import__('datetime').timedelta(days=last_day.day)
                d2 = min(dt_date.fromisoformat(leave["date_to"]), month_end)
                if d2 >= d1:
                    leave_days += (d2 - d1).days + 1
            except Exception:
                pass

        avg_duration = 0
        durations = [r["duration_minutes"] for r in user_records if r.get("duration_minutes")]
        if durations:
            avg_duration = round(sum(durations) / len(durations))

        summary.append({
            "user_id":           user["id"],
            "user_name":         user["name"],
            "user_email":        user["email"],
            "role":              user["role"],
            "days_present":      days_present,
            "days_clocked_in_only": days_clocked_in_only,
            "days_on_leave":     leave_days,
            "avg_duration_minutes": avg_duration,
            "records":           user_records,
        })

    return {"month": month, "summary": summary}


@api_router.get("/attendance/staff/{user_id}")
async def get_staff_attendance(
    user_id: str,
    month: Optional[str] = None,
    current_user: User = Depends(get_current_user)
):
    """Admin/Manager: get a specific staff member's attendance records."""
    require_roles(current_user, ["admin", "manager"],
                  "Only admins and managers can view staff attendance")

    query = {"user_id": user_id}
    if month:
        query["date"] = {"$regex": f"^{month}"}

    records = await db.attendance.find(query, {"_id": 0}).sort("date", -1).to_list(500)
    for r in records:
        if isinstance(r.get("created_at"), str):
            r["created_at"] = datetime.fromisoformat(r["created_at"])
    return records


# ══════════════════════════════════════════════════════════════
#  LEAVE ENDPOINTS
# ══════════════════════════════════════════════════════════════

@api_router.post("/leaves")
async def request_leave(
    leave_data: LeaveRequestCreate,
    current_user: User = Depends(get_current_user)
):
    """Any staff can submit a leave request."""
    if leave_data.leave_type not in ("sick", "casual", "emergency", "other"):
        raise HTTPException(status_code=400, detail="Invalid leave type")
    if not leave_data.reason.strip():
        raise HTTPException(status_code=400, detail="Reason is required")

    # Check for overlapping pending/approved leave
    existing = await db.leaves.find_one({
        "user_id": current_user.id,
        "status":  {"$in": ["pending", "approved"]},
        "date_from": {"$lte": leave_data.date_to},
        "date_to":   {"$gte": leave_data.date_from},
    }, {"_id": 0})
    if existing:
        raise HTTPException(
            status_code=400,
            detail="You already have a leave request overlapping these dates"
        )

    days = _count_working_days(leave_data.date_from, leave_data.date_to)

    leave = LeaveRequest(
        user_id=current_user.id,
        user_name=current_user.name,
        date_from=leave_data.date_from,
        date_to=leave_data.date_to,
        days_count=days,
        leave_type=leave_data.leave_type,
        reason=leave_data.reason.strip(),
    )
    doc = leave.model_dump()
    doc["created_at"] = doc["created_at"].isoformat()
    await db.leaves.insert_one(doc)
    return leave


@api_router.get("/leaves")
async def get_leaves(
    status: Optional[str] = None,
    month:  Optional[str] = None,
    current_user: User = Depends(get_current_user)
):
    """Admin/Manager: all leaves. Staff: own leaves only."""
    query = {}
    if current_user.role not in ["admin", "manager"]:
        query["user_id"] = current_user.id
    if status:
        query["status"] = status
    if month:
        query["date_from"] = {"$regex": f"^{month}"}

    leaves = await db.leaves.find(query, {"_id": 0}).sort("created_at", -1).to_list(1000)
    for l in leaves:
        if isinstance(l.get("created_at"), str):
            l["created_at"] = datetime.fromisoformat(l["created_at"])
    return leaves


@api_router.put("/leaves/{leave_id}/approve")
async def approve_leave(leave_id: str, current_user: User = Depends(get_current_user)):
    require_roles(current_user, ["admin", "manager"], "Only admins/managers can approve leaves")

    leave = await db.leaves.find_one({"id": leave_id}, {"_id": 0})
    if not leave:
        raise HTTPException(status_code=404, detail="Leave request not found")
    if leave["status"] != "pending":
        raise HTTPException(status_code=400, detail=f"Leave is already {leave['status']}")

    now = _now_iso()
    await db.leaves.update_one({"id": leave_id}, {"$set": {
        "status":           "approved",
        "reviewed_by_id":   current_user.id,
        "reviewed_by_name": current_user.name,
        "reviewed_at":      now,
    }})

    # Create attendance records for the approved leave days
    from datetime import date as dt_date, timedelta
    d = dt_date.fromisoformat(leave["date_from"])
    end = dt_date.fromisoformat(leave["date_to"])
    while d <= end:
        date_str = d.isoformat()
        existing = await db.attendance.find_one(
            {"user_id": leave["user_id"], "date": date_str}, {"_id": 0}
        )
        if not existing:
            record = AttendanceRecord(
                user_id=leave["user_id"],
                user_name=leave["user_name"],
                date=date_str,
                status="on_leave",
                notes=f"Approved {leave['leave_type']} leave",
            )
            rdoc = record.model_dump()
            rdoc["created_at"] = rdoc["created_at"].isoformat()
            await db.attendance.insert_one(rdoc)
        d += timedelta(days=1)

    return {"message": "Leave approved"}


@api_router.put("/leaves/{leave_id}/reject")
async def reject_leave(leave_id: str, current_user: User = Depends(get_current_user)):
    require_roles(current_user, ["admin", "manager"], "Only admins/managers can reject leaves")

    leave = await db.leaves.find_one({"id": leave_id}, {"_id": 0})
    if not leave:
        raise HTTPException(status_code=404, detail="Leave request not found")
    if leave["status"] != "pending":
        raise HTTPException(status_code=400, detail=f"Leave is already {leave['status']}")

    await db.leaves.update_one({"id": leave_id}, {"$set": {
        "status":           "rejected",
        "reviewed_by_id":   current_user.id,
        "reviewed_by_name": current_user.name,
        "reviewed_at":      _now_iso(),
    }})
    return {"message": "Leave rejected"}


@api_router.delete("/leaves/{leave_id}")
async def cancel_leave(leave_id: str, current_user: User = Depends(get_current_user)):
    """Staff can cancel their own pending leave. Admin can cancel any."""
    leave = await db.leaves.find_one({"id": leave_id}, {"_id": 0})
    if not leave:
        raise HTTPException(status_code=404, detail="Leave request not found")
    if current_user.role not in ["admin", "manager"] and leave["user_id"] != current_user.id:
        raise HTTPException(status_code=403, detail="Not authorized")
    if leave["status"] == "approved":
        raise HTTPException(status_code=400, detail="Cannot cancel an approved leave. Contact admin.")

    await db.leaves.delete_one({"id": leave_id})
    return {"message": "Leave request cancelled"}

# ── Helper: raise 403 if role not allowed ─────────────────────
def _require_commission(current_user, allowed_roles):
    """Works with Pydantic User model (.role attribute)."""
    if current_user.role not in allowed_roles:
        raise HTTPException(status_code=403, detail="Access denied")
 
# ── Stats / Summary  (MUST be before /{commission_id} routes) ─
@api_router.get("/commissions/stats/summary")
async def commission_stats(
    month: Optional[str] = None,
    current_user: User = Depends(get_current_user)
):
    try:
        query = {}
        if current_user.role not in ["admin", "manager"]:
            query["staff_id"] = current_user.id
        if month:
            query["date"] = {"$regex": f"^{month}"}
 
        all_c = await db.commissions.find(query).to_list(10000)
 
        total_pending = sum(c["amount"] for c in all_c if c.get("status") == "pending")
        total_paid    = sum(c["amount"] for c in all_c if c.get("status") == "paid")
 
        staff_breakdown = {}
        if current_user.role in ["admin", "manager"]:
            for c in all_c:
                sid = c["staff_id"]
                if sid not in staff_breakdown:
                    staff_breakdown[sid] = {
                        "staff_id": sid,
                        "staff_name": c.get("staff_name", ""),
                        "staff_role": c.get("staff_role", "staff"),
                        "total_earned": 0, "pending": 0, "paid": 0, "count": 0,
                    }
                staff_breakdown[sid]["count"] += 1
                staff_breakdown[sid]["total_earned"] += c["amount"]
                if c.get("status") == "pending":
                    staff_breakdown[sid]["pending"] += c["amount"]
                else:
                    staff_breakdown[sid]["paid"] += c["amount"]
 
        return {
            "total_pending":   total_pending,
            "total_paid":      total_paid,
            "total_count":     len(all_c),
            "pending_count":   len([c for c in all_c if c.get("status") == "pending"]),
            "staff_breakdown": sorted(list(staff_breakdown.values()), key=lambda x: -x["total_earned"]),
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Stats error: {str(e)}")
 
# ── Bulk Pay  (MUST be before /{commission_id} routes) ────────
@api_router.put("/commissions/bulk/pay")
async def bulk_pay_commissions(
    data: BulkPayRequest,
    current_user: User = Depends(get_current_user)
):
    _require_commission(current_user, ["admin"])
    try:
        now = datetime.utcnow().isoformat()
        result = await db.commissions.update_many(
            {"id": {"$in": data.commission_ids}, "status": "pending"},
            {"$set": {
                "status": "paid",
                "paid_by": current_user.id,
                "paid_by_name": current_user.name,
                "paid_at": now,
            }}
        )
        return {"updated": result.modified_count}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Bulk pay error: {str(e)}")
 
# ── Create Commission ─────────────────────────────────────────
@api_router.post("/commissions")
async def create_commission(
    data: CommissionCreate,
    current_user: User = Depends(get_current_user)
):
    _require_commission(current_user, ["admin", "manager"])
    try:
        staff = await db.users.find_one({"id": data.staff_id})
        if not staff:
            raise HTTPException(status_code=404, detail="Staff member not found")
 
        commission = {
            "id":               str(uuid.uuid4()),
            "staff_id":         data.staff_id,
            "staff_name":       staff.get("name", ""),
            "staff_email":      staff.get("email", ""),
            "staff_role":       staff.get("role", "staff"),
            "amount":           data.amount,
            "commission_type":  data.commission_type,
            "percentage_value": data.percentage_value,
            "order_id":         data.order_id,
            "order_amount":     data.order_amount,
            "notes":            data.notes,
            "date":             data.date,
            "status":           "pending",
            "created_by":       current_user.id,
            "created_by_name":  current_user.name,
            "paid_by":          None,
            "paid_by_name":     None,
            "paid_at":          None,
            "created_at":       datetime.utcnow().isoformat(),
        }
        await db.commissions.insert_one(commission)
        commission.pop("_id", None)
        return commission
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Create commission error: {str(e)}")
 
# ── List Commissions ──────────────────────────────────────────
@api_router.get("/commissions")
async def list_commissions(
    staff_id: Optional[str] = None,
    status:   Optional[str] = None,
    month:    Optional[str] = None,
    current_user: User = Depends(get_current_user)
):
    try:
        query = {}
        if current_user.role not in ["admin", "manager"]:
            query["staff_id"] = current_user.id
        elif staff_id:
            query["staff_id"] = staff_id
        if status:
            query["status"] = status
        if month:
            query["date"] = {"$regex": f"^{month}"}
 
        commissions = await db.commissions.find(query).sort("created_at", -1).to_list(500)
        for c in commissions:
            c.pop("_id", None)
        return commissions
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"List error: {str(e)}")
 
# ── Mark Single as Paid ───────────────────────────────────────
@api_router.put("/commissions/{commission_id}/pay")
async def pay_commission(
    commission_id: str,
    current_user: User = Depends(get_current_user)
):
    _require_commission(current_user, ["admin"])
    try:
        c = await db.commissions.find_one({"id": commission_id})
        if not c:
            raise HTTPException(status_code=404, detail="Commission not found")
        if c.get("status") == "paid":
            raise HTTPException(status_code=400, detail="Already paid")
        now = datetime.utcnow().isoformat()
        await db.commissions.update_one(
            {"id": commission_id},
            {"$set": {
                "status": "paid",
                "paid_by": current_user.id,
                "paid_by_name": current_user.name,
                "paid_at": now,
            }}
        )
        c.update({"status": "paid", "paid_by": current_user.id,
                   "paid_by_name": current_user.name, "paid_at": now})
        c.pop("_id", None)
        return c
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Pay error: {str(e)}")
 
# ── Delete Commission ─────────────────────────────────────────
@api_router.delete("/commissions/{commission_id}")
async def delete_commission(
    commission_id: str,
    current_user: User = Depends(get_current_user)
):
    _require_commission(current_user, ["admin"])
    try:
        result = await db.commissions.delete_one({"id": commission_id})
        if result.deleted_count == 0:
            raise HTTPException(status_code=404, detail="Commission not found")
        return {"deleted": True}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Delete error: {str(e)}")


# ─── App setup ────────────────────────────────────────────────
app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)


@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()

    # ── Helper: raise 403 if role not allowed ─────────────────────
def _require(current_user, allowed_roles):
    if current_user.get("role") not in allowed_roles:
        raise HTTPException(status_code=403, detail="Access denied")
